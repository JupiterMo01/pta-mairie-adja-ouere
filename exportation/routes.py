import io
from flask import render_template, request, redirect, url_for, flash, session, send_file, abort
from flask_login import login_required
from models import db, Programme, Direction, Service, Annee
from exportation import exportation_bp
from utils import get_annee


def _sheet_name(name, used, prefix=''):
    """Retourne un nom de feuille Excel unique (max 31 chars)."""
    base = (prefix + name)[:31]
    candidate = base
    i = 2
    while candidate in used:
        candidate = (base[:28] + f'_{i}')[:31]
        i += 1
    used.add(candidate)
    return candidate


# ── Feuille PTA Global ────────────────────────────────────────────────────────

def _fill_global_sheet(ws, annee):
    """Remplit ws avec le PTA Global complet (même format que l'export du module PTA)."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime as _dt
    import os as _os
    from flask import current_app

    thin = Side(style='thin')
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    lft  = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    f_prog   = PatternFill("solid", fgColor="FFFF99")
    f_proj   = PatternFill("solid", fgColor="FFB6C1")
    f_act    = PatternFill("solid", fgColor="D3D3D3")
    f_tch    = PatternFill("solid", fgColor="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="FF69B4")
    os_fill  = PatternFill("solid", fgColor="FFFACD")
    tot_fill = PatternFill("solid", fgColor="AED6F1")
    NUM_COLS = {4, 5, 6, 7, 8, 9}

    _MOIS = {
        'Janvier': 'Janv', 'Février': 'Fév', 'Mars': 'Mars', 'Avril': 'Avr',
        'Mai': 'Mai', 'Juin': 'Juin', 'Juillet': 'Juil', 'Août': 'Août',
        'Septembre': 'Sept', 'Octobre': 'Oct', 'Novembre': 'Nov', 'Décembre': 'Déc',
    }

    def fper(d, f):
        d = _MOIS.get(d or '', d or ''); f = _MOIS.get(f or '', f or '')
        return f"{d} - {f}" if d and f and d != f else d or f or ''

    def fmt(v): return v if v else 0

    def wr(vals, fill, bold, rn):
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=rn, column=col, value=v)
            c.fill = fill; c.font = Font(bold=bold, size=9)
            c.alignment = lft if col == 2 else ctr; c.border = brd
            if col in NUM_COLS and isinstance(v, (int, float)):
                c.number_format = '#,##0'

    # ── En-tête institutionnel ─────────────────────────────────────────────────
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 8
    ws.row_dimensions[3].height = 22
    ws.merge_cells('A1:D3'); ws.merge_cells('E1:J3'); ws.merge_cells('K1:O3')
    c = ws['E1']
    c.value = "BP 02 Adja-Ouèrè\nTél : +229 01 61 91 96 12\nEmail : contact.adjaouere@mairie.bj"
    c.font = Font(bold=True, size=9)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    try:
        from openpyxl.drawing.image import Image as _XI
        from PIL import Image as _PI
        import io as _io2
        _img = _os.path.join(current_app.root_path, 'static', 'img')
        _H = 52
        for _path, _anchor in [(_os.path.join(_img, 'bandeau.png'), 'A1'),
                                (_os.path.join(_img, 'logo_commune.png'), None)]:
            if not _os.path.exists(_path): continue
            _buf = _io2.BytesIO()
            with _PI.open(_path) as _p: _ow, _oh = _p.size; _p.save(_buf, 'PNG')
            _buf.seek(0); _xi = _XI(_buf); _xi.height = _H
            _xi.width = int(_ow * _H / _oh)
            if _anchor:
                ws.add_image(_xi, _anchor)
            else:
                try:
                    from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
                    from openpyxl.drawing.xdr import XDRPositiveSize2D
                    _a = OneCellAnchor()
                    _a._from = AnchorMarker(col=15, colOff=-int(_xi.width*9525), row=0, rowOff=0)
                    _a.ext = XDRPositiveSize2D(int(_xi.width*9525), int(_H*9525))
                    _xi.anchor = _a; ws.add_image(_xi)
                except Exception: ws.add_image(_xi, 'N1')
    except Exception: pass

    # ── Titre ──────────────────────────────────────────────────────────────────
    ws.merge_cells('A4:O4')
    ws['A4'].value = f"PLAN DE TRAVAIL ANNUEL — Exercice {annee.annee}"
    ws['A4'].font = Font(bold=True, size=14)
    ws['A4'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A4'].fill = PatternFill("solid", fgColor="D6EAF8")
    _med = Side(style='medium')
    for _tc in range(1, 16):
        ws.cell(row=4, column=_tc).border = Border(
            left=_med if _tc==1 else Side(style=None),
            right=_med if _tc==15 else Side(style=None),
            top=_med, bottom=_med)
    ws.row_dimensions[4].height = 28

    if annee.objectif_general:
        ws.merge_cells('A5:O5'); c = ws['A5']
        c.value = f"Objectif général : {annee.objectif_general}"
        c.font = Font(bold=True, italic=True); c.alignment = lft
        start_row = 6
    else:
        start_row = 5

    # ── En-têtes tableau ───────────────────────────────────────────────────────
    hr1 = ['Code', 'Objectifs/Résultats/Actions/Activités/Tâches', 'Imputation budgétaire',
            'Sources de financement (F CFA)', '', '', '', '', '',
            "Période d'exécution", 'Poids (%)',
            'Direction/Unité Resp.', 'Structures associées / Unité Admi. Associée',
            "Mode d'exécution", 'Observations (Contribution ODD…)']
    hr2 = ['', '', '', 'Source de financement', '', '', '', '', 'Total', '', '', '', '', '', '']
    hr3 = ['', '', '', 'Fonds Propres', 'FA', 'FNA', 'PTFs', 'Autres Fonds', '', '', '', '', '', '', '']
    for off, rd in enumerate([hr1, hr2, hr3]):
        rn = start_row + off
        for col, val in enumerate(rd, 1):
            c = ws.cell(row=rn, column=col, value=val)
            c.fill = hdr_fill; c.font = Font(bold=True, color="000000", size=8)
            c.alignment = ctr; c.border = brd
    ws.merge_cells(f'A{start_row}:A{start_row+2}'); ws.merge_cells(f'B{start_row}:B{start_row+2}')
    ws.merge_cells(f'C{start_row}:C{start_row+2}'); ws.merge_cells(f'D{start_row}:I{start_row}')
    ws.merge_cells(f'D{start_row+1}:H{start_row+1}'); ws.merge_cells(f'I{start_row+1}:I{start_row+2}')
    ws.merge_cells(f'J{start_row}:J{start_row+2}'); ws.merge_cells(f'K{start_row}:K{start_row+2}')
    ws.merge_cells(f'L{start_row}:L{start_row+2}'); ws.merge_cells(f'M{start_row}:M{start_row+2}')
    ws.merge_cells(f'N{start_row}:N{start_row+2}'); ws.merge_cells(f'O{start_row}:O{start_row+2}')

    row = start_row + 3
    programmes = Programme.query.filter_by(annee_id=annee.id).order_by(Programme.numero).all()
    _nb_svc = Service.query.count()
    _nb_dir = Direction.query.count()

    def _assoc(svcs, dirs, ext, acteurs):
        nb_s, nb_d = len(svcs), len(dirs)
        if nb_s == _nb_svc > 0 and nb_d == _nb_dir > 0: return 'Tous services et directions'
        if nb_s == _nb_svc > 0: return 'Tous les services'
        if nb_d == _nb_dir > 0: return 'Toutes les directions'
        parts = [s.code for s in svcs] + [d.code for d in dirs] + [se.nom for se in ext]
        if acteurs: parts.append(acteurs)
        return ', '.join(parts) if parts else 'Aucun'

    tg = [0, 0, 0, 0, 0]
    for prog in programmes:
        if prog.objectif_specifique:
            ws.merge_cells(f'A{row}:O{row}')
            c = ws.cell(row=row, column=1,
                        value=f"Objectif {prog.numero}/Résultat {prog.numero} : {prog.objectif_specifique}")
            c.font = Font(bold=True, italic=True, size=9)
            c.fill = os_fill; c.alignment = lft; row += 1

        p_rp = sum(t.ressources_propres or 0 for pj in prog.projets for a in pj.activites for t in a.taches)
        p_fa = sum(t.fadec_affecte      or 0 for pj in prog.projets for a in pj.activites for t in a.taches)
        p_fn = sum(t.fadec_non_affecte  or 0 for pj in prog.projets for a in pj.activites for t in a.taches)
        p_ap = sum(t.autres_partenaires or 0 for pj in prog.projets for a in pj.activites for t in a.taches)
        p_af = sum(t.autres_fonds       or 0 for pj in prog.projets for a in pj.activites for t in a.taches)
        wr([prog.code, prog.libelle, '',
            fmt(p_rp), fmt(p_fa), fmt(p_fn), fmt(p_ap), fmt(p_af), fmt(p_rp+p_fa+p_fn+p_ap+p_af),
            '', prog.poids, '', '', '', ''], f_prog, True, row); row += 1

        for projet in prog.projets:
            j_rp = sum(t.ressources_propres or 0 for a in projet.activites for t in a.taches)
            j_fa = sum(t.fadec_affecte      or 0 for a in projet.activites for t in a.taches)
            j_fn = sum(t.fadec_non_affecte  or 0 for a in projet.activites for t in a.taches)
            j_ap = sum(t.autres_partenaires or 0 for a in projet.activites for t in a.taches)
            j_af = sum(t.autres_fonds       or 0 for a in projet.activites for t in a.taches)
            wr([projet.code, projet.libelle, '',
                fmt(j_rp), fmt(j_fa), fmt(j_fn), fmt(j_ap), fmt(j_af), fmt(j_rp+j_fa+j_fn+j_ap+j_af),
                '', projet.poids, '', '', '', ''], f_proj, True, row); row += 1

            for act in projet.activites:
                resp_a = act.direction_responsable.code if act.direction_responsable else ''
                a_rp = sum(t.ressources_propres or 0 for t in act.taches)
                a_fa = sum(t.fadec_affecte      or 0 for t in act.taches)
                a_fn = sum(t.fadec_non_affecte  or 0 for t in act.taches)
                a_ap = sum(t.autres_partenaires or 0 for t in act.taches)
                a_af = sum(t.autres_fonds       or 0 for t in act.taches)
                wr([act.code, act.nom, act.imputation_budgetaire or 'NÉANT',
                    fmt(a_rp), fmt(a_fa), fmt(a_fn), fmt(a_ap), fmt(a_af), fmt(a_rp+a_fa+a_fn+a_ap+a_af),
                    fper(act.periode_debut, act.periode_fin), act.poids, resp_a,
                    _assoc(act.services_intervenants, act.directions_associees,
                           act.structures_externes, act.acteurs_externes),
                    act.mode_execution or 'Direct', act.observations or ''],
                   f_act, True, row); row += 1

                for t in sorted(act.taches, key=lambda x: (x.ordre or 0, x.id)):
                    resp_t = (t.service_responsable.code if t.service_responsable
                              else (t.direction_responsable.code if t.direction_responsable else ''))
                    tg[0] += t.ressources_propres or 0; tg[1] += t.fadec_affecte or 0
                    tg[2] += t.fadec_non_affecte  or 0; tg[3] += t.autres_partenaires or 0
                    tg[4] += t.autres_fonds       or 0
                    wr([t.code, f"  {t.nom}", t.imputation_budgetaire or 'NÉANT',
                        fmt(t.ressources_propres), fmt(t.fadec_affecte),
                        fmt(t.fadec_non_affecte), fmt(t.autres_partenaires),
                        fmt(t.autres_fonds), fmt(t.budget_total),
                        fper(t.periode_debut, t.periode_fin), t.poids, resp_t,
                        _assoc(t.services_concernes, t.directions_associees,
                               t.structures_externes, t.acteurs_externes),
                        t.mode_execution or 'Direct', t.observations or ''],
                       f_tch, False, row); row += 1

    wr(['', 'TOTAL GÉNÉRAL', '', tg[0], tg[1], tg[2], tg[3], tg[4], sum(tg),
        '', 100, '', '', '', ''], tot_fill, True, row)
    for col in range(1, 16):
        ws.cell(row=row, column=col).font = Font(bold=True, color="000000", size=9)

    row += 1
    _date_str = _dt.now().strftime('%d/%m/%Y à %H:%M')
    ws.merge_cells(f'A{row}:G{row}')
    c = ws.cell(row=row, column=1, value=f"Exporté le {_date_str}")
    c.font = Font(bold=True, size=8); c.alignment = lft
    ws.merge_cells(f'H{row}:O{row}')
    c = ws.cell(row=row, column=8,
                value="Direction du Développement Local et de la Planification (DDLP)")
    c.font = Font(bold=True, size=8)
    c.alignment = Alignment(horizontal='right', vertical='center')

    widths = [10, 45, 14, 13, 10, 10, 13, 13, 14, 14, 8, 18, 28, 18, 28]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f'A{start_row + 3}'


# ── Feuille récapitulative ────────────────────────────────────────────────────

def _fill_recap_sheet(ws, annee, directions):
    """Feuille de synthèse : budget et tâches par direction et par programme."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    thin = Side(style='thin')
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    lft  = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    hdr  = PatternFill("solid", fgColor="1F4E79")
    alt1 = PatternFill("solid", fgColor="D6EAF8")
    alt2 = PatternFill("solid", fgColor="EBF5FB")
    tot  = PatternFill("solid", fgColor="FFFF99")

    def wr(vals, fill, bold, rn, aligns=None):
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=rn, column=col, value=v)
            c.fill = fill; c.font = Font(bold=bold, size=9); c.border = brd
            c.alignment = (aligns[col-1] if aligns else (lft if col == 1 else ctr))
            if col >= 3 and isinstance(v, (int, float)):
                c.number_format = '#,##0'

    # Titre
    ws.merge_cells('A1:H1')
    ws['A1'].value = f"RÉCAPITULATIF PTA {annee.annee} — par Direction"
    ws['A1'].font = Font(bold=True, size=13, color="FFFFFF")
    ws['A1'].fill = PatternFill("solid", fgColor="1F4E79")
    ws['A1'].alignment = ctr
    ws.row_dimensions[1].height = 28

    # En-têtes
    heads = ['Direction', 'Code', 'Budget (FCFA)', 'Fds Propres', 'FADEC Affecté',
             'FADEC N-Affecté', 'Partenaires', 'Nb Tâches']
    for col, h in enumerate(heads, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.fill = hdr; c.font = Font(bold=True, color="FFFFFF", size=9)
        c.alignment = ctr; c.border = brd
    ws.row_dimensions[2].height = 20

    from dirpta.routes import _compute_pta_direction
    row = 3
    tg_budget = tg_rp = tg_fa = tg_fn = tg_ap = tg_nb = 0

    for idx, d in enumerate(sorted(directions, key=lambda x: x.nom)):
        data = _compute_pta_direction(annee, d)
        rp = sum(td['tache'].ressources_propres or 0
                 for pd in data for pjd in pd['projets']
                 for ad in pjd['activites'] for td in ad['taches'])
        fa = sum(td['tache'].fadec_affecte or 0
                 for pd in data for pjd in pd['projets']
                 for ad in pjd['activites'] for td in ad['taches'])
        fn = sum(td['tache'].fadec_non_affecte or 0
                 for pd in data for pjd in pd['projets']
                 for ad in pjd['activites'] for td in ad['taches'])
        ap = sum(td['tache'].autres_partenaires or 0
                 for pd in data for pjd in pd['projets']
                 for ad in pjd['activites'] for td in ad['taches'])
        af = sum(td['tache'].autres_fonds or 0
                 for pd in data for pjd in pd['projets']
                 for ad in pjd['activites'] for td in ad['taches'])
        nb = sum(len(ad['taches']) for pd in data for pjd in pd['projets']
                 for ad in pjd['activites'])
        budget = rp + fa + fn + ap + af
        tg_budget += budget; tg_rp += rp; tg_fa += fa; tg_fn += fn; tg_ap += ap; tg_nb += nb
        fill = alt1 if idx % 2 == 0 else alt2
        wr([d.nom, d.code, budget, rp, fa, fn, ap, nb], fill, False, row)
        row += 1

    wr(['TOTAL', '', tg_budget, tg_rp, tg_fa, tg_fn, tg_ap, tg_nb], tot, True, row)
    row += 2

    # Section 2 : par Programme
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'].value = f"RÉCAPITULATIF PTA {annee.annee} — par Programme"
    ws[f'A{row}'].font = Font(bold=True, size=13, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill("solid", fgColor="145A32")
    ws[f'A{row}'].alignment = ctr
    ws.row_dimensions[row].height = 28; row += 1

    heads2 = ['Programme', 'N°', 'Budget (FCFA)', 'Fds Propres', 'FADEC Affecté',
              'FADEC N-Affecté', 'Partenaires', 'Nb Tâches']
    for col, h in enumerate(heads2, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = PatternFill("solid", fgColor="1E8449"); c.font = Font(bold=True, color="FFFFFF", size=9)
        c.alignment = ctr; c.border = brd
    ws.row_dimensions[row].height = 20; row += 1

    programmes = Programme.query.filter_by(annee_id=annee.id).order_by(Programme.numero).all()
    for idx, prog in enumerate(programmes):
        rp = sum(t.ressources_propres or 0 for pj in prog.projets for a in pj.activites for t in a.taches)
        fa = sum(t.fadec_affecte      or 0 for pj in prog.projets for a in pj.activites for t in a.taches)
        fn = sum(t.fadec_non_affecte  or 0 for pj in prog.projets for a in pj.activites for t in a.taches)
        ap = sum(t.autres_partenaires or 0 for pj in prog.projets for a in pj.activites for t in a.taches)
        af = sum(t.autres_fonds       or 0 for pj in prog.projets for a in pj.activites for t in a.taches)
        nb = sum(len(a.taches) for pj in prog.projets for a in pj.activites)
        fill = alt1 if idx % 2 == 0 else alt2
        wr([prog.libelle, str(prog.numero), rp+fa+fn+ap+af, rp, fa, fn, ap, nb], fill, False, row)
        row += 1

    widths2 = [42, 6, 18, 15, 15, 16, 15, 12]
    for i, w in enumerate(widths2, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A3'


# ── Constructeur de classeur multi-feuilles ───────────────────────────────────

def _build_workbook(annee, include_global=False, include_recap=False,
                    directions=None, services=None):
    """
    Construit un classeur Excel multi-feuilles.
    Ordre : [Récap ?] [PTA Global ?] [Direction A, ses services…] [Direction B, ses services…] …
    Les directions sont triées alphabétiquement, les services aussi au sein de chaque direction.
    """
    from openpyxl import Workbook
    from dirpta.routes import _fill_sheet as _fdir, _compute_pta_direction
    from svcpta.routes import _fill_sheet as _fsvc, _compute_pta_service

    wb = Workbook()
    wb.remove(wb.active)
    used = set()

    dirs = sorted(directions or [], key=lambda d: d.nom)
    svcs = services or []

    if include_recap and dirs:
        ws = wb.create_sheet(_sheet_name("Récapitulatif", used))
        _fill_recap_sheet(ws, annee, dirs)

    if include_global:
        ws = wb.create_sheet(_sheet_name("PTA Global", used))
        _fill_global_sheet(ws, annee)

    for direction in dirs:
        data = _compute_pta_direction(annee, direction)
        if data:
            ws = wb.create_sheet(_sheet_name(direction.code, used))
            _fdir(ws, annee, direction, data)

        # Services de cette direction (parmi ceux demandés), triés par nom
        dir_svcs = sorted(
            [s for s in svcs if s.direction_id == direction.id],
            key=lambda s: s.nom
        )
        for service in dir_svcs:
            data_s = _compute_pta_service(annee, service)
            if data_s:
                ws = wb.create_sheet(_sheet_name(service.code, used))
                _fsvc(ws, annee, service, data_s)

    return wb


def _send_wb(wb, filename):
    output = io.BytesIO()
    wb.save(output); output.seek(0)
    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)


# ── Routes ────────────────────────────────────────────────────────────────────

@exportation_bp.route('/')
@login_required
def index():
    annee      = get_annee()
    directions = Direction.query.order_by(Direction.nom).all()
    services   = Service.query.order_by(Service.nom).all()
    programmes = Programme.query.filter_by(annee_id=annee.id).order_by(Programme.numero).all() if annee else []
    return render_template('exportation/index.html',
                           annee=annee, directions=directions,
                           services=services, programmes=programmes)


# ── 1. PTA Global : impression / Excel ───────────────────────────────────────

@exportation_bp.route('/print-global')
@login_required
def print_global():
    return redirect(url_for('pta.print_view'))


@exportation_bp.route('/excel-global')
@login_required
def excel_global():
    return redirect(url_for('pta.export_excel'))


# ── 2. Entité unique : direction ou service ───────────────────────────────────

@exportation_bp.route('/action-entite', methods=['POST'])
@login_required
def action_entite():
    """Redirige vers impression ou export Excel d'une direction ou d'un service."""
    etype  = request.form.get('etype')   # 'direction' ou 'service'
    eid    = request.form.get('eid', type=int)
    action = request.form.get('action')  # 'print' ou 'excel'

    if not eid:
        flash('Veuillez sélectionner une direction ou un service.', 'warning')
        return redirect(url_for('exportation.index'))

    if etype == 'direction':
        if action == 'print':
            return redirect(url_for('dirpta.print_view', direction_id=eid))
        return redirect(url_for('dirpta.export_excel', direction_id=eid))
    else:
        if action == 'print':
            return redirect(url_for('svcpta.print_view', service_id=eid))
        return redirect(url_for('svcpta.export_excel', service_id=eid))


# ── 3. Sélection multiple directions et/ou services ──────────────────────────

@exportation_bp.route('/excel-selection', methods=['POST'])
@login_required
def excel_selection():
    annee = get_annee()
    if not annee:
        flash('Aucune année active.', 'danger')
        return redirect(url_for('exportation.index'))

    dir_ids = request.form.getlist('dir_ids', type=int)
    svc_ids = request.form.getlist('svc_ids', type=int)
    avec_recap = 'avec_recap' in request.form

    if not dir_ids and not svc_ids:
        flash('Sélectionnez au moins une direction ou un service.', 'warning')
        return redirect(url_for('exportation.index'))

    dirs = Direction.query.filter(Direction.id.in_(dir_ids)).all() if dir_ids else []
    svcs = Service.query.filter(Service.id.in_(svc_ids)).all() if svc_ids else []

    # Pour les services sélectionnés sans leur direction sélectionnée,
    # on ajoute la direction comme référence de tri (sans générer sa feuille)
    dir_ids_pour_tri = set(dir_ids)
    dirs_de_svcs = {s.direction for s in svcs if s.direction_id not in dir_ids_pour_tri}

    # On construit : directions explicitement sélectionnées + directions "porteuses" de services
    all_dirs_for_build = list(dirs) + list(dirs_de_svcs)

    wb = _build_workbook(annee, include_recap=avec_recap,
                         directions=all_dirs_for_build, services=svcs)

    # Supprimer les feuilles de directions non sélectionnées explicitement
    # (elles ont pu être créées pour porter leurs services)
    dirs_non_selectionnees = {d.code for d in dirs_de_svcs}
    sheets_a_retirer = [ws for ws in wb.worksheets if ws.title in dirs_non_selectionnees]
    for ws in sheets_a_retirer:
        del wb[ws.title]

    if not wb.worksheets:
        flash('Aucune donnée pour les entités sélectionnées.', 'warning')
        return redirect(url_for('exportation.index'))

    return _send_wb(wb, f"PTA_Selection_{annee.annee}.xlsx")


# ── 4. Une direction + tous ses services ──────────────────────────────────────

@exportation_bp.route('/excel-direction-et-services/<int:direction_id>')
@login_required
def excel_direction_et_services(direction_id):
    annee     = get_annee()
    direction = Direction.query.get_or_404(direction_id)
    # Contrôle IDOR : une direction ne voit que sa propre direction ;
    # un service ne voit que la direction qui le contient.
    if current_user.role == 'direction' and current_user.direction_id != direction_id:
        abort(403)
    if current_user.role == 'service' and (
            not current_user.service or
            current_user.service.direction_id != direction_id):
        abort(403)
    if not annee:
        flash('Aucune année active.', 'danger')
        return redirect(url_for('exportation.index'))

    services = Service.query.filter_by(direction_id=direction.id).order_by(Service.nom).all()
    wb = _build_workbook(annee, directions=[direction], services=services)

    if not wb.worksheets:
        flash(f'Aucune donnée pour {direction.code}.', 'warning')
        return redirect(url_for('exportation.index'))

    return _send_wb(wb, f"PTA_{direction.code}_et_services_{annee.annee}.xlsx")


# ── 5. Toutes les directions + leurs services ─────────────────────────────────

@exportation_bp.route('/excel-toutes-dirs-services')
@login_required
def excel_toutes_dirs_services():
    annee = get_annee()
    if not annee:
        flash('Aucune année active.', 'danger')
        return redirect(url_for('exportation.index'))

    directions = Direction.query.order_by(Direction.nom).all()
    services   = Service.query.order_by(Service.nom).all()
    avec_recap = request.args.get('recap', '0') == '1'

    wb = _build_workbook(annee, include_recap=avec_recap,
                         directions=directions, services=services)
    return _send_wb(wb, f"PTA_Toutes_Directions_et_Services_{annee.annee}.xlsx")


# ── 6. Export global complet ──────────────────────────────────────────────────

@exportation_bp.route('/excel-complet')
@login_required
def excel_complet():
    annee = get_annee()
    if not annee:
        flash('Aucune année active.', 'danger')
        return redirect(url_for('exportation.index'))

    directions = Direction.query.order_by(Direction.nom).all()
    services   = Service.query.order_by(Service.nom).all()
    avec_recap = request.args.get('recap', '0') == '1'

    wb = _build_workbook(annee, include_global=True, include_recap=avec_recap,
                         directions=directions, services=services)
    return _send_wb(wb, f"PTA_Complet_{annee.annee}.xlsx")


# ── 7. Export par programme(s) ────────────────────────────────────────────────

@exportation_bp.route('/excel-par-programme', methods=['POST'])
@login_required
def excel_par_programme():
    """Exporte seulement les programmes sélectionnés (PTA global filtré)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime as _dt

    annee = get_annee()
    if not annee:
        flash('Aucune année active.', 'danger')
        return redirect(url_for('exportation.index'))

    prog_ids = request.form.getlist('prog_ids', type=int)
    if not prog_ids:
        flash('Sélectionnez au moins un programme.', 'warning')
        return redirect(url_for('exportation.index'))

    programmes = Programme.query.filter(
        Programme.id.in_(prog_ids), Programme.annee_id == annee.id
    ).order_by(Programme.numero).all()

    if not programmes:
        flash('Aucun programme trouvé.', 'warning')
        return redirect(url_for('exportation.index'))

    wb = Workbook()
    ws = wb.active
    ws.title = f"PTA {annee.annee} — Programmes"

    # On réutilise _fill_global_sheet mais avec programmes filtrés :
    # On remplace temporairement la requête en passant les programmes directement
    # En pratique on crée une version simplifiée
    thin = Side(style='thin')
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    lft  = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    f_prog   = PatternFill("solid", fgColor="FFFF99")
    f_proj   = PatternFill("solid", fgColor="FFB6C1")
    f_act    = PatternFill("solid", fgColor="D3D3D3")
    f_tch    = PatternFill("solid", fgColor="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="FF69B4")
    os_fill  = PatternFill("solid", fgColor="FFFACD")
    tot_fill = PatternFill("solid", fgColor="AED6F1")
    NUM_COLS = {4, 5, 6, 7, 8, 9}
    _MOIS = {'Janvier':'Janv','Février':'Fév','Mars':'Mars','Avril':'Avr','Mai':'Mai',
             'Juin':'Juin','Juillet':'Juil','Août':'Août','Septembre':'Sept',
             'Octobre':'Oct','Novembre':'Nov','Décembre':'Déc'}
    def fper(d, f):
        d=_MOIS.get(d or '',d or ''); f=_MOIS.get(f or '',f or '')
        return f"{d} - {f}" if d and f and d!=f else d or f or ''
    def fmt(v): return v if v else 0
    def wr(vals, fill, bold, rn):
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=rn, column=col, value=v)
            c.fill=fill; c.font=Font(bold=bold, size=9)
            c.alignment=lft if col==2 else ctr; c.border=brd
            if col in NUM_COLS and isinstance(v,(int,float)): c.number_format='#,##0'

    noms_prog = ', '.join(str(p.numero) for p in programmes)
    ws.merge_cells('A1:O1')
    ws['A1'].value = (f"PLAN DE TRAVAIL ANNUEL {annee.annee}"
                      f" — Programme(s) : {noms_prog}")
    ws['A1'].font = Font(bold=True, size=13)
    ws['A1'].alignment = ctr
    ws['A1'].fill = PatternFill("solid", fgColor="D5F5E3")
    ws.row_dimensions[1].height = 26

    start_row = 2
    hr1 = ['Code','Objectifs/Résultats/Actions/Activités/Tâches','Imputation budgétaire',
           'Sources de financement (F CFA)','','','','','',
           "Période d'exécution",'Poids (%)','Direction/Unité Resp.',
           'Structures associées',"Mode d'exécution",'Observations']
    hr2 = ['','','','Source de financement','','','','','Total','','','','','','']
    hr3 = ['','','','Fonds Propres','FA','FNA','PTFs','Autres Fonds','','','','','','','']
    for off, rd in enumerate([hr1, hr2, hr3]):
        rn = start_row + off
        for col, val in enumerate(rd, 1):
            c = ws.cell(row=rn, column=col, value=val)
            c.fill=hdr_fill; c.font=Font(bold=True, color="000000", size=8)
            c.alignment=ctr; c.border=brd
    ws.merge_cells(f'A{start_row}:A{start_row+2}'); ws.merge_cells(f'B{start_row}:B{start_row+2}')
    ws.merge_cells(f'C{start_row}:C{start_row+2}'); ws.merge_cells(f'D{start_row}:I{start_row}')
    ws.merge_cells(f'D{start_row+1}:H{start_row+1}'); ws.merge_cells(f'I{start_row+1}:I{start_row+2}')
    for col_l in 'JKLMNO':
        ws.merge_cells(f'{col_l}{start_row}:{col_l}{start_row+2}')

    row = start_row + 3
    _nb_svc = Service.query.count(); _nb_dir = Direction.query.count()
    def _assoc(svcs, dirs, ext, act):
        nb_s,nb_d=len(svcs),len(dirs)
        if nb_s==_nb_svc>0 and nb_d==_nb_dir>0: return 'Tous services et directions'
        if nb_s==_nb_svc>0: return 'Tous les services'
        if nb_d==_nb_dir>0: return 'Toutes les directions'
        parts=[s.code for s in svcs]+[d.code for d in dirs]+[se.nom for se in ext]
        if act: parts.append(act)
        return ', '.join(parts) if parts else 'Aucun'

    tg=[0,0,0,0,0]
    for prog in programmes:
        if prog.objectif_specifique:
            ws.merge_cells(f'A{row}:O{row}')
            c=ws.cell(row=row,column=1,
                      value=f"Objectif {prog.numero}/Résultat {prog.numero} : {prog.objectif_specifique}")
            c.font=Font(bold=True,italic=True,size=9); c.fill=os_fill; c.alignment=lft; row+=1
        p_rp=sum(t.ressources_propres or 0 for pj in prog.projets for a in pj.activites for t in a.taches)
        p_fa=sum(t.fadec_affecte or 0 for pj in prog.projets for a in pj.activites for t in a.taches)
        p_fn=sum(t.fadec_non_affecte or 0 for pj in prog.projets for a in pj.activites for t in a.taches)
        p_ap=sum(t.autres_partenaires or 0 for pj in prog.projets for a in pj.activites for t in a.taches)
        p_af=sum(t.autres_fonds or 0 for pj in prog.projets for a in pj.activites for t in a.taches)
        wr([prog.code,prog.libelle,'',fmt(p_rp),fmt(p_fa),fmt(p_fn),fmt(p_ap),fmt(p_af),
            fmt(p_rp+p_fa+p_fn+p_ap+p_af),'',prog.poids,'','','',''],f_prog,True,row); row+=1
        for projet in prog.projets:
            j_rp=sum(t.ressources_propres or 0 for a in projet.activites for t in a.taches)
            j_fa=sum(t.fadec_affecte or 0 for a in projet.activites for t in a.taches)
            j_fn=sum(t.fadec_non_affecte or 0 for a in projet.activites for t in a.taches)
            j_ap=sum(t.autres_partenaires or 0 for a in projet.activites for t in a.taches)
            j_af=sum(t.autres_fonds or 0 for a in projet.activites for t in a.taches)
            wr([projet.code,projet.libelle,'',fmt(j_rp),fmt(j_fa),fmt(j_fn),fmt(j_ap),fmt(j_af),
                fmt(j_rp+j_fa+j_fn+j_ap+j_af),'',projet.poids,'','','',''],f_proj,True,row); row+=1
            for act in projet.activites:
                resp_a=act.direction_responsable.code if act.direction_responsable else ''
                a_rp=sum(t.ressources_propres or 0 for t in act.taches)
                a_fa=sum(t.fadec_affecte or 0 for t in act.taches)
                a_fn=sum(t.fadec_non_affecte or 0 for t in act.taches)
                a_ap=sum(t.autres_partenaires or 0 for t in act.taches)
                a_af=sum(t.autres_fonds or 0 for t in act.taches)
                wr([act.code,act.nom,act.imputation_budgetaire or 'NÉANT',
                    fmt(a_rp),fmt(a_fa),fmt(a_fn),fmt(a_ap),fmt(a_af),
                    fmt(a_rp+a_fa+a_fn+a_ap+a_af),
                    fper(act.periode_debut,act.periode_fin),act.poids,resp_a,
                    _assoc(act.services_intervenants,act.directions_associees,
                           act.structures_externes,act.acteurs_externes),
                    act.mode_execution or 'Direct',act.observations or ''],f_act,True,row); row+=1
                for t in sorted(act.taches, key=lambda x:(x.ordre or 0,x.id)):
                    resp_t=(t.service_responsable.code if t.service_responsable
                            else (t.direction_responsable.code if t.direction_responsable else ''))
                    tg[0]+=t.ressources_propres or 0; tg[1]+=t.fadec_affecte or 0
                    tg[2]+=t.fadec_non_affecte or 0; tg[3]+=t.autres_partenaires or 0
                    tg[4]+=t.autres_fonds or 0
                    wr([t.code,f"  {t.nom}",t.imputation_budgetaire or 'NÉANT',
                        fmt(t.ressources_propres),fmt(t.fadec_affecte),
                        fmt(t.fadec_non_affecte),fmt(t.autres_partenaires),
                        fmt(t.autres_fonds),fmt(t.budget_total),
                        fper(t.periode_debut,t.periode_fin),t.poids,resp_t,
                        _assoc(t.services_concernes,t.directions_associees,
                               t.structures_externes,t.acteurs_externes),
                        t.mode_execution or 'Direct',t.observations or ''],f_tch,False,row); row+=1

    wr(['','TOTAL GÉNÉRAL','',tg[0],tg[1],tg[2],tg[3],tg[4],sum(tg),'',100,'','','',''],
       tot_fill,True,row)
    for col in range(1,16):
        ws.cell(row=row,column=col).font=Font(bold=True,color="000000",size=9)

    widths=[10,45,14,13,10,10,13,13,14,14,8,18,28,18,28]
    for i,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(i)].width=w
    ws.freeze_panes=f'A{start_row+3}'

    noms_fichier = '_'.join(str(p.numero) for p in programmes)
    return _send_wb(wb, f"PTA_{annee.annee}_Prog_{noms_fichier}.xlsx")