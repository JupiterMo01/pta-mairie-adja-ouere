import io
from flask import render_template, redirect, url_for, flash, request, session, send_file
from flask_login import login_required, current_user
from models import Annee, Direction, Service
from rapports import rapports_bp
from utils import rapport_trimestriel


def get_annee():
    annee_id = session.get('annee_id')
    if annee_id:
        return Annee.query.get(annee_id)
    return Annee.query.filter_by(actif=True).first()


def _get_filtres():
    """Détermine les filtres service/direction selon le rôle de l'utilisateur."""
    service_id = None
    direction_id = None
    if current_user.role == 'service':
        service_id = current_user.service_id
    elif current_user.role == 'direction':
        direction_id = current_user.direction_id
    elif current_user.role in ('admin_editeur', 'admin_lecteur'):
        service_id = request.args.get('service_id', type=int)
        direction_id = request.args.get('direction_id', type=int)
    return service_id, direction_id


@rapports_bp.route('/')
@login_required
def index():
    annee = get_annee()
    trimestre = request.args.get('trimestre', 1, type=int)
    service_id, direction_id = _get_filtres()

    data, global_taux = [], None
    if annee:
        data, global_taux = rapport_trimestriel(annee, trimestre, service_id, direction_id)

    services = Service.query.order_by(Service.nom).all()
    directions = Direction.query.order_by(Direction.nom).all()

    return render_template('rapports/index.html',
                           data=data, annee=annee, trimestre=trimestre,
                           global_taux=global_taux, trimestres=[1, 2, 3, 4],
                           services=services, directions=directions,
                           filtre_service_id=service_id,
                           filtre_direction_id=direction_id)


@rapports_bp.route('/export/excel/<int:trimestre>')
@login_required
def export_rapport_excel(trimestre):
    annee = get_annee()
    if not annee:
        flash('Aucune année active.', 'danger')
        return redirect(url_for('rapports.index'))

    service_id, direction_id = _get_filtres()
    data, global_taux = rapport_trimestriel(annee, trimestre, service_id, direction_id)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = f"Rapport T{trimestre} {annee.annee}"

    thin = Side(style='thin')
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr = Alignment(horizontal='center', vertical='center', wrap_text=True)
    lft = Alignment(horizontal='left', vertical='center', wrap_text=True)

    fills = {
        'titre': PatternFill("solid", fgColor="1F4E79"),
        'prog':  PatternFill("solid", fgColor="1F4E79"),
        'proj':  PatternFill("solid", fgColor="2E75B6"),
        'act':   PatternFill("solid", fgColor="9DC3E6"),
        'tch':   PatternFill("solid", fgColor="DEEAF1"),
        'hdr':   PatternFill("solid", fgColor="2E75B6"),
    }

    # Titre
    ws.merge_cells('A1:F1')
    ws['A1'] = f"RAPPORT D'ÉVALUATION DU PTA — TRIMESTRE {trimestre} — {annee.annee}"
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['A1'].fill = fills['titre']
    ws['A1'].alignment = ctr

    ws.merge_cells('A2:F2')
    taux_txt = f"{global_taux:.2f} %" if global_taux is not None else "N/A"
    ws['A2'] = f"Taux global d'exécution : {taux_txt}"
    ws['A2'].font = Font(bold=True, size=12)
    ws['A2'].alignment = ctr

    headers = ['Code', 'Désignation', 'Poids (%)', 'Statut', 'Taux (%)', 'Observation']
    ws.append(headers)
    for col, _ in enumerate(headers, 1):
        c = ws.cell(row=3, column=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = fills['hdr']
        c.alignment = ctr
        c.border = brd

    statut_labels = {'execute': 'Exécuté', 'en_cours': 'En cours', 'non_execute': 'Non exécuté'}
    row = 4

    for pd in data:
        prog = pd['programme']
        ws.append([prog.code, prog.nom, round(prog.poids or 0, 2), '',
                   f"{pd['taux']:.2f} %" if pd['taux'] is not None else 'N/A', ''])
        for col in range(1, 7):
            c = ws.cell(row=row, column=col)
            c.fill = fills['prog']
            c.font = Font(bold=True, color="FFFFFF")
            c.alignment = lft if col == 2 else ctr
            c.border = brd
        row += 1

        for pjd in pd['projets']:
            proj = pjd['projet']
            ws.append([proj.code, proj.nom, round(proj.poids or 0, 2), '',
                       f"{pjd['taux']:.2f} %" if pjd['taux'] is not None else 'N/A', ''])
            for col in range(1, 7):
                c = ws.cell(row=row, column=col)
                c.fill = fills['proj']
                c.font = Font(bold=True, color="FFFFFF")
                c.alignment = lft if col == 2 else ctr
                c.border = brd
            row += 1

            for ad in pjd['activites']:
                act = ad['activite']
                ws.append([act.code, act.nom, round(act.poids or 0, 2), '',
                           f"{ad['taux']:.2f} %" if ad['taux'] is not None else 'N/A', ''])
                for col in range(1, 7):
                    c = ws.cell(row=row, column=col)
                    c.fill = fills['act']
                    c.font = Font(bold=True)
                    c.alignment = lft if col == 2 else ctr
                    c.border = brd
                row += 1

                for td in ad['taches']:
                    tache = td['tache']
                    suivi = td['suivi']
                    statut = suivi.statut if suivi else 'non_execute'
                    obs = suivi.observation if suivi else ''
                    ws.append([
                        tache.code, f"  {tache.nom}", round(tache.poids or 0, 2),
                        statut_labels.get(statut, 'Non exécuté'),
                        f"{td['taux']:.0f} %", obs or '',
                    ])
                    for col in range(1, 7):
                        c = ws.cell(row=row, column=col)
                        c.fill = fills['tch']
                        c.alignment = lft if col in (2, 4, 6) else ctr
                        c.border = brd
                        if col == 5:
                            if td['taux'] == 100:
                                c.font = Font(bold=True, color="375623")
                            elif td['taux'] == 0:
                                c.font = Font(color="9C0006")
                    row += 1

    for i, w in enumerate([15, 50, 12, 18, 12, 40], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A4'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    fname = f"Rapport_T{trimestre}_{annee.annee}.xlsx"
    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=fname)