import io
import json
from functools import wraps
from flask import render_template, redirect, url_for, flash, request, session, send_file, jsonify
from flask_login import login_required, current_user
from sqlalchemy.orm import subqueryload
from models import db, Programme, Projet, Activite, Tache, Direction, Service, Annee, MODES_EXECUTION, BiblioActivite, BiblioTache, StructureExterne, PTABackup
from pta import pta_bp
from utils import get_annee


def editeur_only(f):
    @wraps(f)
    @login_required
    def decorated(*args, **kwargs):
        if current_user.role != 'admin_editeur':
            flash('Accès refusé. Réservé aux éditeurs.', 'danger')
            return redirect(url_for('pta.global_pta'))
        return f(*args, **kwargs)
    return decorated


def _parse_float(val, default=0.0):
    try:
        if isinstance(val, str):
            val = val.strip().replace(' ', '').replace(' ', '')
            if ',' in val:
                val = val.replace('.', '').replace(',', '.')
        return float(val or 0)
    except (ValueError, TypeError):
        return default


def _update_associations(obj, services_ids, directions_ids, services_rel, directions_rel):
    """Met à jour les relations services et directions associés."""
    setattr(obj, services_rel, [])
    for sid in services_ids:
        s = Service.query.get(int(sid))
        if s:
            getattr(obj, services_rel).append(s)
    setattr(obj, directions_rel, [])
    for did in directions_ids:
        d = Direction.query.get(int(did))
        if d:
            getattr(obj, directions_rel).append(d)


def _renumeroter_taches(activite_id):
    """Renuméroter les tâches selon leur ordre."""
    taches = Tache.query.filter_by(activite_id=activite_id)\
                        .order_by(Tache.ordre, Tache.numero).all()
    for i, t in enumerate(taches, 1):
        t.numero = i
        t.ordre = i
    db.session.flush()


# ─── Vue principale ──────────────────────────────────────────────────────────

@pta_bp.route('/')
@login_required
def global_pta():
    annee = get_annee()
    annees = Annee.query.order_by(Annee.annee.desc()).all()
    programmes = []
    if annee:
        programmes = Programme.query.filter_by(annee_id=annee.id)\
                                    .order_by(Programme.numero)\
                                    .options(
                                        subqueryload(Programme.projets)
                                        .subqueryload(Projet.activites)
                                        .subqueryload(Activite.taches)
                                    ).all()
    directions = Direction.query.order_by(Direction.nom).all()
    services = Service.query.order_by(Service.nom).all()
    can_edit = (current_user.role == 'admin_editeur')
    mon_service_id = current_user.service_id if current_user.role == 'service' else None
    ma_direction_id = current_user.direction_id if current_user.role == 'direction' else None

    biblio_activites = BiblioActivite.query.order_by(BiblioActivite.nom).all()
    structures_externes = StructureExterne.query.order_by(StructureExterne.nom).all()
    return render_template('pta/global.html',
                           programmes=programmes, annee=annee, annees=annees,
                           directions=directions, services=services,
                           can_edit=can_edit,
                           modes_execution=MODES_EXECUTION,
                           mon_service_id=mon_service_id,
                           ma_direction_id=ma_direction_id,
                           biblio_activites=biblio_activites,
                           structures_externes=structures_externes)


@pta_bp.route('/annee/<int:annee_id>')
@login_required
def switch_annee(annee_id):
    annee = Annee.query.get_or_404(annee_id)
    session['annee_id'] = annee.id
    session['annee'] = annee.annee
    return redirect(url_for('pta.global_pta'))


# ─── Objectif général de l'année ─────────────────────────────────────────────

@pta_bp.route('/annee/<int:annee_id>/objectif', methods=['POST'])
@editeur_only
def annee_objectif(annee_id):
    annee = Annee.query.get_or_404(annee_id)
    annee.objectif_general = request.form.get('objectif_general', '').strip()
    db.session.commit()
    flash('Objectif général mis à jour.', 'success')
    return redirect(url_for('pta.global_pta'))


# ─── Programmes ──────────────────────────────────────────────────────────────

@pta_bp.route('/programme/add', methods=['POST'])
@editeur_only
def programme_add():
    annee = get_annee()
    if not annee:
        flash('Aucune année active.', 'danger')
        return redirect(url_for('pta.global_pta'))
    nom = request.form.get('nom', '').strip()
    poids = _parse_float(request.form.get('poids'))
    objectif_specifique = request.form.get('objectif_specifique', '').strip()
    if not nom:
        flash('Le nom est obligatoire.', 'danger')
        return redirect(url_for('pta.global_pta'))
    insert_after = request.form.get('insert_after', type=int)
    if insert_after is not None:
        for pp in Programme.query.filter_by(annee_id=annee.id)\
                                  .filter(Programme.numero > insert_after).all():
            pp.numero += 1
        db.session.flush()
        numero = insert_after + 1
    else:
        last = Programme.query.filter_by(annee_id=annee.id).order_by(Programme.numero.desc()).first()
        numero = (last.numero + 1) if last else 1
    p = Programme(annee_id=annee.id, numero=numero, nom=nom,
                  poids=poids, objectif_specifique=objectif_specifique)
    db.session.add(p)
    db.session.commit()
    flash(f'Programme {numero} ajouté.', 'success')
    return redirect(url_for('pta.global_pta', go=f'prog-{p.id}'))


@pta_bp.route('/programme/<int:prog_id>/edit', methods=['POST'])
@editeur_only
def programme_edit(prog_id):
    p = Programme.query.get_or_404(prog_id)
    p.nom = request.form.get('nom', '').strip()
    p.poids = _parse_float(request.form.get('poids'))
    p.objectif_specifique = request.form.get('objectif_specifique', '').strip()
    p.description = request.form.get('description', '').strip()
    db.session.commit()
    flash('Programme mis à jour.', 'success')
    return redirect(url_for('pta.global_pta'))


@pta_bp.route('/programme/<int:prog_id>/delete', methods=['POST'])
@editeur_only
def programme_delete(prog_id):
    p = Programme.query.get_or_404(prog_id)
    annee_id = p.annee_id
    db.session.delete(p)
    db.session.flush()
    _renumeroter_programmes(annee_id)
    db.session.commit()
    flash('Programme supprimé.', 'success')
    return redirect(url_for('pta.global_pta'))


# ─── Projets ─────────────────────────────────────────────────────────────────

@pta_bp.route('/projet/add', methods=['POST'])
@editeur_only
def projet_add():
    prog_id = int(request.form.get('programme_id'))
    programme = Programme.query.get_or_404(prog_id)
    nom = request.form.get('nom', '').strip()
    poids = _parse_float(request.form.get('poids'))
    last = Projet.query.filter_by(programme_id=prog_id).order_by(Projet.numero.desc()).first()
    numero = (last.numero + 1) if last else 1
    db.session.add(Projet(programme_id=prog_id, numero=numero, nom=nom, poids=poids,
                          description=request.form.get('description', '').strip()))
    db.session.commit()
    flash(f'Projet {programme.numero}.{numero} ajouté.', 'success')
    return redirect(url_for('pta.global_pta'))


@pta_bp.route('/projet/<int:proj_id>/edit', methods=['POST'])
@editeur_only
def projet_edit(proj_id):
    p = Projet.query.get_or_404(proj_id)
    p.nom = request.form.get('nom', '').strip()
    p.poids = _parse_float(request.form.get('poids'))
    p.description = request.form.get('description', '').strip()
    db.session.commit()
    flash('Projet mis à jour.', 'success')
    return redirect(url_for('pta.global_pta'))


@pta_bp.route('/projet/<int:proj_id>/delete', methods=['POST'])
@editeur_only
def projet_delete(proj_id):
    p = Projet.query.get_or_404(proj_id)
    prog_id = p.programme_id
    db.session.delete(p)
    db.session.flush()
    _renumeroter_projets(prog_id)
    db.session.commit()
    flash('Projet supprimé.', 'success')
    return redirect(url_for('pta.global_pta', go=f'prog-{prog_id}'))


@pta_bp.route('/projet/<int:proj_id>/clear_activites', methods=['POST'])
@editeur_only
def projet_clear_activites(proj_id):
    p = Projet.query.get_or_404(proj_id)
    code = p.code
    for a in list(p.activites):
        db.session.delete(a)
    db.session.commit()
    flash(f'Toutes les activités du projet {code} supprimées.', 'warning')
    return redirect(url_for('pta.global_pta', go=f'projet-{proj_id}'))


# ─── Activités ───────────────────────────────────────────────────────────────

def _activite_form_overrides(a):
    """Applique seulement les valeurs du formulaire qui sont explicitement renseignées.
    Utilisée après un import (source-first) pour laisser l'utilisateur écraser."""
    nom = request.form.get('nom', '').strip()
    if nom:
        a.nom = nom
    desc = request.form.get('description', '').strip()
    if desc:
        a.description = desc
    pd = request.form.get('periode_debut', '').strip()
    if pd:
        a.periode_debut = pd
    pf = request.form.get('periode_fin', '').strip()
    if pf:
        a.periode_fin = pf
    dir_id = request.form.get('direction_responsable_id', '').strip()
    if dir_id:
        a.direction_responsable_id = int(dir_id)
    imp_type = request.form.get('imputation_type', 'neant')
    if imp_type == 'ligne':
        imp_text = request.form.get('imputation_budgetaire', '').strip()
        if imp_text:
            a.imputation_budgetaire = imp_text
    details = request.form.get('details_financement', '').strip()
    if details and details.upper() != 'RAS':
        a.details_financement = details
    acteurs = request.form.get('acteurs_externes', '').strip()
    if acteurs:
        a.acteurs_externes = acteurs
    obs = request.form.get('observations', '').strip()
    a.observations = obs or None
    mode = request.form.get('mode_execution', '').strip()
    if mode and mode != 'Direct':
        a.mode_execution = mode
    poids_str = request.form.get('poids', '').strip()
    if poids_str:
        p = _parse_float(poids_str)
        if p:
            a.poids = p
    structures_mode = request.form.get('structures_mode', 'aucun')
    if structures_mode == 'tous_services':
        a.services_intervenants = Service.query.order_by(Service.nom).all()
        a.directions_associees = []
    elif structures_mode == 'toutes_directions':
        a.services_intervenants = []
        a.directions_associees = Direction.query.order_by(Direction.nom).all()
    elif structures_mode == 'tous_sd':
        a.services_intervenants = Service.query.order_by(Service.nom).all()
        a.directions_associees = Direction.query.order_by(Direction.nom).all()
    elif structures_mode == 'manuel':
        _update_associations(a,
                             request.form.getlist('services_intervenants'),
                             request.form.getlist('directions_associees'),
                             'services_intervenants', 'directions_associees')
    se_ids = request.form.getlist('structures_externes')
    if se_ids:
        a.structures_externes = [StructureExterne.query.get(int(i)) for i in se_ids if i]


def _activite_from_form(a):
    """Remplit les champs d'une activité depuis le formulaire."""
    a.nom = request.form.get('nom', '').strip()
    a.poids = _parse_float(request.form.get('poids'))
    a.description = request.form.get('description', '').strip()
    a.direction_responsable_id = request.form.get('direction_responsable_id') or None
    if a.direction_responsable_id:
        a.direction_responsable_id = int(a.direction_responsable_id)
    imp_type = request.form.get('imputation_type', 'neant')
    if imp_type == 'ligne':
        a.imputation_budgetaire = request.form.get('imputation_budgetaire', '').strip() or 'NÉANT'
    else:
        a.imputation_budgetaire = 'NÉANT'
    a.periode_debut = request.form.get('periode_debut', '').strip()
    a.periode_fin = request.form.get('periode_fin', '').strip()
    a.mode_execution = request.form.get('mode_execution', '').strip() or 'Direct'
    a.type_activite = request.form.get('type_activite', 'Activité de fonctionnement').strip() or 'Activité de fonctionnement'
    a.details_financement = request.form.get('details_financement', '').strip() or 'RAS'
    a.acteurs_externes = request.form.get('acteurs_externes', '').strip()
    a.observations = request.form.get('observations', '').strip() or None
    structures_mode = request.form.get('structures_mode', 'aucun')
    if structures_mode == 'tous_services':
        a.services_intervenants = Service.query.order_by(Service.nom).all()
        a.directions_associees = []
    elif structures_mode == 'toutes_directions':
        a.services_intervenants = []
        a.directions_associees = Direction.query.order_by(Direction.nom).all()
    elif structures_mode == 'tous_sd':
        a.services_intervenants = Service.query.order_by(Service.nom).all()
        a.directions_associees = Direction.query.order_by(Direction.nom).all()
    elif structures_mode == 'manuel':
        _update_associations(a,
                             request.form.getlist('services_intervenants'),
                             request.form.getlist('directions_associees'),
                             'services_intervenants', 'directions_associees')
    else:  # aucun
        a.services_intervenants = []
        a.directions_associees = []
    a.structures_externes = [StructureExterne.query.get(int(i))
                             for i in request.form.getlist('structures_externes') if i]


@pta_bp.route('/activite/add', methods=['POST'])
@editeur_only
def activite_add():
    try:
        proj_id = int(request.form.get('projet_id', 0))
    except (ValueError, TypeError):
        flash('Paramètre projet invalide.', 'danger')
        return redirect(url_for('pta.global_pta'))
    Projet.query.get_or_404(proj_id)
    insert_after = request.form.get('insert_after', type=int)
    if insert_after is not None:
        acts_apres = Activite.query.filter_by(projet_id=proj_id)\
                                    .filter(Activite.numero > insert_after).all()
        for aa in acts_apres:
            aa.numero += 1
        db.session.flush()
        numero = insert_after + 1
    else:
        last = Activite.query.filter_by(projet_id=proj_id).order_by(Activite.numero.desc()).first()
        numero = (last.numero + 1) if last else 1
    import_source = request.form.get('import_source', '')
    import_id_val = request.form.get('import_id', type=int)

    a = Activite(projet_id=proj_id, numero=numero)

    if import_source == 'biblio' and import_id_val:
        ba = BiblioActivite.query.get(import_id_val)
        if ba:
            # 1. Charger toutes les données depuis la bibliothèque
            a.nom = ba.nom or ''
            a.description = ba.description or ''
            a.periode_debut = ba.periode_debut or ''
            a.periode_fin = ba.periode_fin or ''
            a.direction_responsable_id = ba.direction_responsable_id
            a.services_intervenants = list(ba.services_associes)
            a.directions_associees = list(ba.directions_associees)
            a.acteurs_externes = ba.acteurs_externes or ''
            a.details_financement = ba.details_financement or 'RAS'
            a.imputation_budgetaire = ba.imputation_budgetaire or 'NÉANT'
            a.ressources_propres = ba.ressources_propres or 0
            a.fadec_affecte = ba.fadec_affecte or 0
            a.fadec_non_affecte = ba.fadec_non_affecte or 0
            a.autres_partenaires = ba.autres_partenaires or 0
            a.autres_fonds = ba.autres_fonds or 0
            a.poids = ba.poids or 0
            a.mode_execution = ba.mode_execution or 'Direct'
            a.type_activite = ba.type_activite or 'Activité de fonctionnement'
            a.structures_externes = list(ba.structures_externes)
            # 2. Laisser le formulaire écraser seulement les champs explicitement renseignés
            _activite_form_overrides(a)
            db.session.add(a)
            db.session.flush()
            _copier_taches_biblio(ba, a)
        else:
            _activite_from_form(a)
            db.session.add(a)
            db.session.flush()
    elif import_source == 'pta' and import_id_val:
        src = Activite.query.get(import_id_val)
        if src:
            a.nom = src.nom or ''
            a.description = src.description or ''
            a.periode_debut = src.periode_debut or ''
            a.periode_fin = src.periode_fin or ''
            a.direction_responsable_id = src.direction_responsable_id
            a.services_intervenants = list(src.services_intervenants)
            a.directions_associees = list(src.directions_associees)
            a.acteurs_externes = src.acteurs_externes or ''
            a.details_financement = src.details_financement or 'RAS'
            a.imputation_budgetaire = src.imputation_budgetaire or 'NÉANT'
            a.ressources_propres = src.ressources_propres or 0
            a.fadec_affecte = src.fadec_affecte or 0
            a.fadec_non_affecte = src.fadec_non_affecte or 0
            a.autres_partenaires = src.autres_partenaires or 0
            a.autres_fonds = src.autres_fonds or 0
            a.poids = src.poids or 0
            a.mode_execution = src.mode_execution or 'Direct'
            a.structures_externes = list(src.structures_externes)
            _activite_form_overrides(a)
            db.session.add(a)
            db.session.flush()
            _copier_taches(src, a)
        else:
            _activite_from_form(a)
            db.session.add(a)
            db.session.flush()
    else:
        _activite_from_form(a)
        db.session.add(a)
        db.session.flush()

    _renumeroter_activites(proj_id)
    db.session.commit()
    flash(f'Activité {a.code} ajoutée.', 'success')
    return redirect(url_for('pta.global_pta', go=f'act-{a.id}'))


@pta_bp.route('/activite/<int:act_id>/edit', methods=['POST'])
@editeur_only
def activite_edit(act_id):
    a = Activite.query.get_or_404(act_id)
    old_imputation = a.imputation_budgetaire
    _activite_from_form(a)
    # Si imputation passe à NÉANT, forcer toutes les tâches à NÉANT
    if a.imputation_budgetaire == 'NÉANT' and old_imputation != 'NÉANT':
        for t in a.taches:
            t.imputation_budgetaire = 'NÉANT'
    db.session.commit()
    flash('Activité mise à jour.', 'success')
    return redirect(url_for('pta.global_pta', go=f'act-{act_id}'))


@pta_bp.route('/activite/<int:act_id>/delete', methods=['POST'])
@editeur_only
def activite_delete(act_id):
    a = Activite.query.get_or_404(act_id)
    proj_id = a.projet_id
    db.session.delete(a)
    db.session.flush()
    _renumeroter_activites(proj_id)
    db.session.commit()
    flash('Activité supprimée.', 'success')
    return redirect(url_for('pta.global_pta', go=f'projet-{proj_id}'))


@pta_bp.route('/activite/<int:act_id>/clear_taches', methods=['POST'])
@editeur_only
def activite_clear_taches(act_id):
    a = Activite.query.get_or_404(act_id)
    code = a.code
    for t in list(a.taches):
        db.session.delete(t)
    db.session.commit()
    flash(f'Toutes les tâches de {code} supprimées.', 'warning')
    return redirect(url_for('pta.global_pta', go=f'act-{act_id}'))


def _renumeroter_activites(projet_id):
    acts = Activite.query.filter_by(projet_id=projet_id).order_by(Activite.numero).all()
    for i, a in enumerate(acts, 1):
        a.numero = i
    db.session.flush()


def _renumeroter_projets(programme_id):
    projets = Projet.query.filter_by(programme_id=programme_id).order_by(Projet.numero).all()
    for i, p in enumerate(projets, 1):
        p.numero = i
    db.session.flush()


def _renumeroter_programmes(annee_id):
    programmes = Programme.query.filter_by(annee_id=annee_id).order_by(Programme.numero).all()
    for i, p in enumerate(programmes, 1):
        p.numero = i
    db.session.flush()


def _copier_taches_biblio(biblio_act, dest_act):
    """Copie les tâches depuis la bibliothèque en conservant tous les champs."""
    for bt in sorted(biblio_act.taches, key=lambda x: x.numero):
        nt = Tache(
            activite_id=dest_act.id,
            numero=bt.numero, ordre=bt.numero, nom=bt.nom,
            description=bt.description, poids=bt.poids,
            service_responsable_id=bt.service_responsable_id,
            direction_responsable_id=bt.direction_responsable_id,
            imputation_budgetaire=bt.imputation_budgetaire or 'NÉANT',
            mode_execution=bt.mode_execution or 'Direct',
            periode_debut=bt.periode_debut,
            periode_fin=bt.periode_fin,
            ressources_propres=bt.ressources_propres or 0,
            fadec_affecte=bt.fadec_affecte or 0,
            fadec_non_affecte=bt.fadec_non_affecte or 0,
            autres_partenaires=bt.autres_partenaires or 0,
            autres_fonds=bt.autres_fonds or 0,
            details_financement=bt.details_financement,
            acteurs_externes=bt.acteurs_externes,
        )
        nt.services_concernes = list(bt.services_concernes)
        nt.directions_associees = list(bt.directions_associees)
        nt.structures_externes = list(bt.structures_externes)
        db.session.add(nt)


def _copier_taches(source_act, dest_act):
    for t in source_act.taches:
        nt = Tache(
            activite_id=dest_act.id,
            numero=t.numero, ordre=t.ordre, nom=t.nom,
            description=t.description, poids=t.poids,
            service_responsable_id=t.service_responsable_id,
            direction_responsable_id=t.direction_responsable_id,
            imputation_budgetaire=t.imputation_budgetaire,
            mode_execution=t.mode_execution or 'Direct',
            ressources_propres=t.ressources_propres or 0,
            fadec_affecte=t.fadec_affecte or 0,
            fadec_non_affecte=t.fadec_non_affecte or 0,
            autres_partenaires=t.autres_partenaires or 0,
            autres_fonds=t.autres_fonds or 0,
            details_financement=t.details_financement,
            acteurs_externes=t.acteurs_externes,
        )
        nt.services_concernes = list(t.services_concernes)
        nt.directions_associees = list(t.directions_associees)
        nt.structures_externes = list(t.structures_externes)
        db.session.add(nt)


@pta_bp.route('/activite/<int:act_id>/move/<direction>', methods=['POST'])
@editeur_only
def activite_move(act_id, direction):
    a = Activite.query.get_or_404(act_id)
    proj_id = a.projet_id
    acts = Activite.query.filter_by(projet_id=proj_id).order_by(Activite.numero).all()
    idx = next((i for i, x in enumerate(acts) if x.id == act_id), None)
    if direction == 'up' and idx and idx > 0:
        acts[idx].numero, acts[idx - 1].numero = acts[idx - 1].numero, acts[idx].numero
    elif direction == 'down' and idx is not None and idx < len(acts) - 1:
        acts[idx].numero, acts[idx + 1].numero = acts[idx + 1].numero, acts[idx].numero
    db.session.flush()
    _renumeroter_activites(proj_id)
    db.session.commit()
    return redirect(url_for('pta.global_pta', go=f'act-{act_id}'))


@pta_bp.route('/activite/<int:act_id>/duplicate', methods=['POST'])
@editeur_only
def activite_duplicate(act_id):
    a = Activite.query.get_or_404(act_id)
    proj_id = a.projet_id
    placement = request.form.get('placement', 'end')
    if placement == 'after':
        acts_apres = Activite.query.filter_by(projet_id=proj_id)\
                                    .filter(Activite.numero > a.numero).all()
        for aa in acts_apres:
            aa.numero += 1
        db.session.flush()
        new_num = a.numero + 1
    else:
        last = Activite.query.filter_by(projet_id=proj_id).order_by(Activite.numero.desc()).first()
        new_num = (last.numero + 1) if last else 1
    na = Activite(
        projet_id=proj_id, numero=new_num,
        nom=a.nom, description=a.description, poids=a.poids,
        direction_responsable_id=a.direction_responsable_id,
        imputation_budgetaire=a.imputation_budgetaire,
        periode_debut=a.periode_debut, periode_fin=a.periode_fin,
        mode_execution=a.mode_execution or 'Direct',
        type_activite=a.type_activite or 'Activité de fonctionnement',
        details_financement=a.details_financement,
        acteurs_externes=a.acteurs_externes,
        ressources_propres=a.ressources_propres or 0,
        fadec_affecte=a.fadec_affecte or 0,
        fadec_non_affecte=a.fadec_non_affecte or 0,
        autres_partenaires=a.autres_partenaires or 0,
        autres_fonds=a.autres_fonds or 0,
    )
    na.services_intervenants = list(a.services_intervenants)
    na.directions_associees = list(a.directions_associees)
    na.structures_externes = list(a.structures_externes)
    db.session.add(na)
    db.session.flush()
    _copier_taches(a, na)
    _renumeroter_activites(proj_id)
    db.session.commit()
    flash(f'Activité « {a.nom} » dupliquée.', 'success')
    return redirect(url_for('pta.global_pta', go=f'act-{na.id}'))


@pta_bp.route('/activite/<int:act_id>/export_biblio', methods=['POST'])
@editeur_only
def activite_export_biblio(act_id):
    a = Activite.query.get_or_404(act_id)
    nom = request.form.get('nom', '').strip() or a.nom
    ba = BiblioActivite(
        nom=nom,
        description=a.description,
        direction_responsable_id=a.direction_responsable_id,
        mode_execution=a.mode_execution or 'Direct',
        type_activite=a.type_activite or 'Activité de fonctionnement',
        imputation_budgetaire=a.imputation_budgetaire,
        periode_debut=a.periode_debut,
        periode_fin=a.periode_fin,
        poids=a.poids,
        ressources_propres=a.ressources_propres or 0,
        fadec_affecte=a.fadec_affecte or 0,
        fadec_non_affecte=a.fadec_non_affecte or 0,
        autres_partenaires=a.autres_partenaires or 0,
        autres_fonds=a.autres_fonds or 0,
        details_financement=a.details_financement,
        acteurs_externes=a.acteurs_externes,
    )
    ba.services_associes = list(a.services_intervenants)
    ba.directions_associees = list(a.directions_associees)
    ba.structures_externes = list(a.structures_externes)
    db.session.add(ba)
    db.session.flush()
    for t in sorted(a.taches, key=lambda x: x.numero):
        bt = BiblioTache(
            biblio_activite_id=ba.id,
            numero=t.numero, nom=t.nom,
            description=t.description, poids=t.poids,
            service_responsable_id=t.service_responsable_id,
            direction_responsable_id=t.direction_responsable_id,
            imputation_budgetaire=t.imputation_budgetaire or 'NÉANT',
            mode_execution=t.mode_execution or 'Direct',
            periode_debut=t.periode_debut,
            periode_fin=t.periode_fin,
            ressources_propres=t.ressources_propres or 0,
            fadec_affecte=t.fadec_affecte or 0,
            fadec_non_affecte=t.fadec_non_affecte or 0,
            autres_partenaires=t.autres_partenaires or 0,
            autres_fonds=t.autres_fonds or 0,
            details_financement=t.details_financement,
            acteurs_externes=t.acteurs_externes,
        )
        bt.services_concernes = list(t.services_concernes)
        bt.directions_associees = list(t.directions_associees)
        bt.structures_externes = list(t.structures_externes)
        db.session.add(bt)
    db.session.commit()
    flash(f'Activité « {nom} » et ses {len(a.taches)} tâche(s) exportées vers la bibliothèque — tous les champs conservés.', 'success')
    return redirect(url_for('pta.global_pta', go=f'act-{act_id}'))


@pta_bp.route('/activite/<int:act_id>/copy_to', methods=['POST'])
@editeur_only
def activite_copy_to(act_id):
    a = Activite.query.get_or_404(act_id)
    target_proj_id = int(request.form.get('projet_id'))
    Projet.query.get_or_404(target_proj_id)
    last = Activite.query.filter_by(projet_id=target_proj_id).order_by(Activite.numero.desc()).first()
    new_num = (last.numero + 1) if last else 1
    na = Activite(
        projet_id=target_proj_id, numero=new_num,
        nom=a.nom, description=a.description, poids=a.poids,
        direction_responsable_id=a.direction_responsable_id,
        imputation_budgetaire=a.imputation_budgetaire,
        periode_debut=a.periode_debut, periode_fin=a.periode_fin,
        mode_execution=a.mode_execution or 'Direct',
        type_activite=a.type_activite or 'Activité de fonctionnement',
        details_financement=a.details_financement,
        acteurs_externes=a.acteurs_externes,
        ressources_propres=a.ressources_propres or 0,
        fadec_affecte=a.fadec_affecte or 0,
        fadec_non_affecte=a.fadec_non_affecte or 0,
        autres_partenaires=a.autres_partenaires or 0,
        autres_fonds=a.autres_fonds or 0,
    )
    na.services_intervenants = list(a.services_intervenants)
    na.directions_associees = list(a.directions_associees)
    na.structures_externes = list(a.structures_externes)
    db.session.add(na)
    db.session.flush()
    _copier_taches(a, na)
    _renumeroter_activites(target_proj_id)
    db.session.commit()
    flash(f'Activité « {a.nom} » copiée dans le projet cible.', 'success')
    return redirect(url_for('pta.global_pta', go=f'act-{na.id}'))


# ─── Tâches ──────────────────────────────────────────────────────────────────

def _tache_from_form(t, activite):
    t.nom = request.form.get('nom', '').strip()
    t.poids = _parse_float(request.form.get('poids'))
    t.description = request.form.get('description', '').strip()
    # Responsable : service OU direction (préfixe s_ ou d_)
    resp_ref = request.form.get('responsable_ref', '').strip()
    if resp_ref.startswith('s_'):
        t.service_responsable_id = int(resp_ref[2:])
        t.direction_responsable_id = None
    elif resp_ref.startswith('d_'):
        t.direction_responsable_id = int(resp_ref[2:])
        t.service_responsable_id = None
    else:
        t.service_responsable_id = None
        t.direction_responsable_id = None
    # Imputation budgétaire contrainte par l'activité
    if activite.imputation_budgetaire == 'NÉANT':
        t.imputation_budgetaire = 'NÉANT'
    else:
        t.imputation_budgetaire = request.form.get('imputation_budgetaire', '').strip() or 'NÉANT'
    # Sources de financement
    t.ressources_propres = _parse_float(request.form.get('ressources_propres'))
    t.fadec_affecte = _parse_float(request.form.get('fadec_affecte'))
    t.fadec_non_affecte = _parse_float(request.form.get('fadec_non_affecte'))
    t.autres_partenaires = _parse_float(request.form.get('autres_partenaires'))
    t.autres_fonds = _parse_float(request.form.get('autres_fonds'))
    t.details_financement = request.form.get('details_financement', '').strip() or 'RAS'
    t.acteurs_externes = request.form.get('acteurs_externes', '').strip()
    t.observations = request.form.get('observations', '').strip() or None
    t.mode_execution = request.form.get('mode_execution', '').strip() or 'Direct'
    t.periode_debut = request.form.get('periode_debut', '').strip() or None
    t.periode_fin = request.form.get('periode_fin', '').strip() or None
    structures_mode = request.form.get('structures_mode', 'aucun')
    if structures_mode == 'tous_services':
        t.services_concernes = Service.query.order_by(Service.nom).all()
        t.directions_associees = []
    elif structures_mode == 'toutes_directions':
        t.services_concernes = []
        t.directions_associees = Direction.query.order_by(Direction.nom).all()
    elif structures_mode == 'tous_sd':
        t.services_concernes = Service.query.order_by(Service.nom).all()
        t.directions_associees = Direction.query.order_by(Direction.nom).all()
    elif structures_mode == 'manuel':
        _update_associations(t,
                             request.form.getlist('services_concernes'),
                             request.form.getlist('directions_associees'),
                             'services_concernes', 'directions_associees')
    else:  # aucun
        t.services_concernes = []
        t.directions_associees = []
    t.structures_externes = [StructureExterne.query.get(int(i))
                             for i in request.form.getlist('structures_externes') if i]


@pta_bp.route('/tache/add', methods=['POST'])
@editeur_only
def tache_add():
    try:
        act_id = int(request.form.get('activite_id', 0))
    except (ValueError, TypeError):
        flash('Paramètre activité invalide.', 'danger')
        return redirect(url_for('pta.global_pta'))
    activite = Activite.query.get_or_404(act_id)
    # Insertion à une position donnée ou en fin
    insert_after = request.form.get('insert_after', type=int)

    if insert_after is not None:
        # Décaler les tâches après la position d'insertion
        taches = Tache.query.filter_by(activite_id=act_id)\
                             .order_by(Tache.ordre).all()
        for t in taches:
            if t.ordre > insert_after:
                t.ordre += 1
        new_ordre = insert_after + 1
    else:
        last = Tache.query.filter_by(activite_id=act_id).order_by(Tache.ordre.desc()).first()
        new_ordre = (last.ordre + 1) if last else 1

    t = Tache(activite_id=act_id, numero=new_ordre, ordre=new_ordre)
    _tache_from_form(t, activite)
    db.session.add(t)
    db.session.flush()
    _renumeroter_taches(act_id)
    db.session.commit()
    flash('Tâche ajoutée.', 'success')
    return redirect(url_for('pta.global_pta', go=f'tache-{t.id}'))


@pta_bp.route('/tache/<int:tache_id>/duplicate', methods=['POST'])
@editeur_only
def tache_duplicate(tache_id):
    t = Tache.query.get_or_404(tache_id)
    act_id = t.activite_id
    placement = request.form.get('placement', 'end')
    if placement == 'after':
        taches_apres = Tache.query.filter_by(activite_id=act_id)\
                                   .filter(Tache.ordre > t.ordre).all()
        for ta in taches_apres:
            ta.ordre += 1
        db.session.flush()
        new_ordre = t.ordre + 1
    else:
        last = Tache.query.filter_by(activite_id=act_id).order_by(Tache.ordre.desc()).first()
        new_ordre = (last.ordre + 1) if last else 1
    nt = Tache(
        activite_id=act_id, numero=new_ordre, ordre=new_ordre,
        nom=t.nom, description=t.description, poids=t.poids,
        service_responsable_id=t.service_responsable_id,
        direction_responsable_id=t.direction_responsable_id,
        imputation_budgetaire=t.imputation_budgetaire,
        mode_execution=t.mode_execution or 'Direct',
        ressources_propres=t.ressources_propres or 0,
        fadec_affecte=t.fadec_affecte or 0,
        fadec_non_affecte=t.fadec_non_affecte or 0,
        autres_partenaires=t.autres_partenaires or 0,
        autres_fonds=t.autres_fonds or 0,
        details_financement=t.details_financement,
        acteurs_externes=t.acteurs_externes,
        observations=t.observations,
        periode_debut=t.periode_debut,
        periode_fin=t.periode_fin,
    )
    nt.services_concernes = list(t.services_concernes)
    nt.directions_associees = list(t.directions_associees)
    nt.structures_externes = list(t.structures_externes)
    db.session.add(nt)
    db.session.flush()
    _renumeroter_taches(act_id)
    db.session.commit()
    flash(f'Tâche « {t.nom} » dupliquée.', 'success')
    return redirect(url_for('pta.global_pta', go=f'tache-{nt.id}'))


@pta_bp.route('/tache/<int:tache_id>/edit', methods=['POST'])
@editeur_only
def tache_edit(tache_id):
    t = Tache.query.get_or_404(tache_id)
    activite = t.activite
    _tache_from_form(t, activite)
    db.session.commit()
    flash('Tâche mise à jour.', 'success')
    return redirect(url_for('pta.global_pta', go=f'tache-{tache_id}'))


@pta_bp.route('/tache/<int:tache_id>/delete', methods=['POST'])
@editeur_only
def tache_delete(tache_id):
    t = Tache.query.get_or_404(tache_id)
    act_id = t.activite_id
    db.session.delete(t)
    db.session.flush()
    _renumeroter_taches(act_id)
    db.session.commit()
    flash('Tâche supprimée.', 'success')
    return redirect(url_for('pta.global_pta', go=f'act-{act_id}'))


@pta_bp.route('/tache/<int:tache_id>/move/<direction>', methods=['POST'])
@editeur_only
def tache_move(tache_id, direction):
    """Déplace une tâche vers le haut ou vers le bas."""
    t = Tache.query.get_or_404(tache_id)
    act_id = t.activite_id
    taches = Tache.query.filter_by(activite_id=act_id).order_by(Tache.ordre).all()
    idx = next((i for i, x in enumerate(taches) if x.id == tache_id), None)

    if direction == 'up' and idx and idx > 0:
        taches[idx].ordre, taches[idx - 1].ordre = taches[idx - 1].ordre, taches[idx].ordre
    elif direction == 'down' and idx is not None and idx < len(taches) - 1:
        taches[idx].ordre, taches[idx + 1].ordre = taches[idx + 1].ordre, taches[idx].ordre

    db.session.flush()
    _renumeroter_taches(act_id)
    db.session.commit()
    return redirect(url_for('pta.global_pta', go=f'tache-{tache_id}'))


# ─── Modals AJAX (chargés à la demande, évite de tout embarquer dans la page) ──

@pta_bp.route('/tache/<int:tache_id>/modal')
@editeur_only
def tache_modal_edit(tache_id):
    """Retourne le contenu HTML du modal d'édition d'une tâche (AJAX)."""
    tache = Tache.query.get_or_404(tache_id)
    return render_template('pta/_modal_edit_tache.html',
        tache=tache,
        services=Service.query.order_by(Service.nom).all(),
        directions=Direction.query.order_by(Direction.nom).all(),
        structures_externes=StructureExterne.query.order_by(StructureExterne.nom).all(),
        modes_execution=MODES_EXECUTION)


@pta_bp.route('/activite/<int:act_id>/modal-add-tache')
@editeur_only
def activite_modal_add_tache(act_id):
    """Retourne le contenu HTML du modal d'ajout d'une tâche (AJAX)."""
    activite = Activite.query.get_or_404(act_id)
    insert_after = request.args.get('insert_after', '')
    return render_template('pta/_modal_add_tache.html',
        activite=activite,
        insert_after=insert_after,
        tache=None,
        services=Service.query.order_by(Service.nom).all(),
        directions=Direction.query.order_by(Direction.nom).all(),
        structures_externes=StructureExterne.query.order_by(StructureExterne.nom).all(),
        modes_execution=MODES_EXECUTION)


@pta_bp.route('/activite/<int:act_id>/modal')
@editeur_only
def activite_modal_edit(act_id):
    """Retourne le contenu HTML du modal d'édition d'une activité (AJAX)."""
    act = Activite.query.get_or_404(act_id)
    annee = get_annee()
    programmes = []
    if annee:
        programmes = Programme.query.filter_by(annee_id=annee.id)\
                                    .order_by(Programme.numero)\
                                    .options(subqueryload(Programme.projets)).all()
    return render_template('pta/_modal_edit_activite.html',
        act=act,
        projet=act.projet,
        services=Service.query.order_by(Service.nom).all(),
        directions=Direction.query.order_by(Direction.nom).all(),
        structures_externes=StructureExterne.query.order_by(StructureExterne.nom).all(),
        biblio_activites=BiblioActivite.query.order_by(BiblioActivite.nom).all(),
        programmes=programmes,
        modes_execution=MODES_EXECUTION)


@pta_bp.route('/projet/<int:proj_id>/modal-add-activite')
@editeur_only
def projet_modal_add_activite(proj_id):
    """Retourne le contenu HTML du modal d'ajout d'une activité (AJAX)."""
    projet = Projet.query.get_or_404(proj_id)
    insert_after = request.args.get('insert_after', '')
    annee = get_annee()
    programmes = []
    if annee:
        programmes = Programme.query.filter_by(annee_id=annee.id)\
                                    .order_by(Programme.numero)\
                                    .options(subqueryload(Programme.projets)).all()
    return render_template('pta/_modal_add_activite.html',
        projet=projet,
        act=None,
        insert_after=insert_after,
        services=Service.query.order_by(Service.nom).all(),
        directions=Direction.query.order_by(Direction.nom).all(),
        structures_externes=StructureExterne.query.order_by(StructureExterne.nom).all(),
        biblio_activites=BiblioActivite.query.order_by(BiblioActivite.nom).all(),
        programmes=programmes,
        modes_execution=MODES_EXECUTION)


@pta_bp.route('/activite/<int:act_id>/modal-copy')
@editeur_only
def activite_modal_copy(act_id):
    """Retourne le contenu HTML du modal de copie d'activité (AJAX)."""
    act = Activite.query.get_or_404(act_id)
    annee = get_annee()
    programmes = []
    if annee:
        programmes = Programme.query.filter_by(annee_id=annee.id)\
                                    .order_by(Programme.numero)\
                                    .options(subqueryload(Programme.projets)).all()
    return render_template('pta/_modal_copy_activite.html',
        act=act, programmes=programmes)


@pta_bp.route('/activite/<int:act_id>/modal-biblio')
@editeur_only
def activite_modal_biblio(act_id):
    """Retourne le contenu HTML du modal d'export bibliothèque (AJAX)."""
    act = Activite.query.get_or_404(act_id)
    return render_template('pta/_modal_biblio_activite.html', act=act)


# ─── Impression ───────────────────────────────────────────────────────────────

@pta_bp.route('/print')
@login_required
def print_view():
    annee = get_annee()
    if not annee:
        flash('Aucune année active.', 'danger')
        return redirect(url_for('pta.global_pta'))
    service_filtre = current_user.service_id if current_user.role == 'service' else None
    nature_filtre  = request.args.get('nature', '').strip()   # 'inv', 'fct' ou ''
    programmes = Programme.query.filter_by(annee_id=annee.id).order_by(Programme.numero).all()
    nb_services = Service.query.count()
    nb_directions = Direction.query.count()
    return render_template('pta/print_view.html',
                           annee=annee,
                           programmes=programmes,
                           service_filtre=service_filtre,
                           nature_filtre=nature_filtre,
                           nb_services=nb_services,
                           nb_directions=nb_directions)


# ─── Export Excel ─────────────────────────────────────────────────────────────

@pta_bp.route('/export/excel')
@login_required
def export_excel():
    annee = get_annee()
    if not annee:
        flash('Aucune année active.', 'danger')
        return redirect(url_for('pta.global_pta'))

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    _nature_label = {'inv': ' - Investissement', 'fct': ' - Fonctionnement'}.get(
        request.args.get('nature', '').strip(), '')
    ws.title = f"PTA {annee.annee}{_nature_label}"[:31]

    thin = Side(style='thin')
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr = Alignment(horizontal='center', vertical='center', wrap_text=True)
    lft = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Couleurs : programme=jaune clair, projet=rose clair, activité=gris clair, tâche=blanc
    f_prog = PatternFill("solid", fgColor="FFFF99")   # jaune clair
    f_proj = PatternFill("solid", fgColor="FFB6C1")   # rose clair
    f_act  = PatternFill("solid", fgColor="D3D3D3")   # gris clair
    f_tch  = PatternFill("solid", fgColor="FFFFFF")   # blanc

    service_filtre = current_user.service_id if current_user.role == 'service' else None
    nature_filtre  = request.args.get('nature', '').strip()   # 'inv', 'fct' ou ''

    import os as _os
    from flask import current_app
    _static_img = _os.path.join(current_app.root_path, 'static', 'img')

    # ── Lignes 1-3 : En-tête institutionnel ──────────────────────────────────
    # 3 zones fusionnées sur 3 lignes, proportions ~31 / 30 / 39 %
    # Gauche  A1:D3 : bandeau.png  (armoirie + MAIRIE + tricolore + REPUBLIQUE)
    # Centre  E1:J3 : contacts (BP, Tél, Email)
    # Droite  K1:O3 : logo commune

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 8
    ws.row_dimensions[3].height = 22

    # Gauche : zone bandeau (cellules vides — image flottante)
    ws.merge_cells('A1:D3')

    # Centre : contacts
    ws.merge_cells('E1:J3')
    c = ws['E1']
    c.value = "BP 02 Adja-Ouèrè\nTél : +229 01 61 91 96 12\nEmail : contact.adjaouere@mairie.bj"
    c.font = Font(bold=True, size=9)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Droite : zone logo (cellules vides — image flottante)
    ws.merge_cells('K1:O3')

    # Images — PIL charge en mémoire (évite les erreurs d'encodage Windows)
    try:
        from openpyxl.drawing.image import Image as _XLImg
        from PIL import Image as _PILImg
        import io as _imgio

        _band_path = _os.path.join(_static_img, 'bandeau.png')
        _logo_path = _os.path.join(_static_img, 'logo_commune.png')

        _H = 52   # hauteur cible pour les deux images (pt/px)

        if _os.path.exists(_band_path):
            _buf = _imgio.BytesIO()
            with _PILImg.open(_band_path) as _pil:
                _ow, _oh = _pil.size
                _pil.save(_buf, format='PNG')
            _buf.seek(0)
            _i = _XLImg(_buf)
            _i.height = _H
            _i.width = int(_ow * _H / _oh)   # largeur proportionnelle
            ws.add_image(_i, 'A1')

        if _os.path.exists(_logo_path):
            _buf2 = _imgio.BytesIO()
            with _PILImg.open(_logo_path) as _pil2:
                _ow2, _oh2 = _pil2.size
                _pil2.save(_buf2, format='PNG')
            _buf2.seek(0)
            _i2 = _XLImg(_buf2)
            _lw = int(_ow2 * _H / _oh2)   # largeur proportionnelle
            _lh = _H
            _i2.height = _lh
            _i2.width = _lw
            # Flush-right via ancrage relatif à la colonne :
            # col=15 = bord DROIT de la colonne O (la 15e, index 14)
            # colOff négatif = décale vers la gauche de exactement la largeur du logo
            # → bord droit logo = bord droit tableau, sans calcul pixel approximatif
            try:
                from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
                from openpyxl.drawing.xdr import XDRPositiveSize2D
                _lw_emu = int(_lw * 9525)
                _lh_emu = int(_lh * 9525)
                _anch2 = OneCellAnchor()
                _anch2._from = AnchorMarker(col=15, colOff=-_lw_emu, row=0, rowOff=0)
                _anch2.ext = XDRPositiveSize2D(_lw_emu, _lh_emu)
                _i2.anchor = _anch2
                ws.add_image(_i2)
            except Exception:
                ws.add_image(_i2, 'N1')
    except Exception:
        pass

    # ── Ligne 4 : Titre PTA (cadre complet) ──────────────────────────────────
    ws.merge_cells('A4:O4')
    ws['A4'].value = f"PLAN DE TRAVAIL ANNUEL — Exercice {annee.annee}"
    ws['A4'].font = Font(bold=True, size=14)
    ws['A4'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A4'].fill = PatternFill("solid", fgColor="D6EAF8")
    _med = Side(style='medium')
    # Cadre sur toutes les cellules de la ligne fusionnée pour un encadrement complet
    for _tc in range(1, 16):
        _cell = ws.cell(row=4, column=_tc)
        _cell.border = Border(
            left=_med  if _tc == 1  else Side(style=None),
            right=_med if _tc == 15 else Side(style=None),
            top=_med, bottom=_med
        )
    ws.row_dimensions[4].height = 28

    # ── Ligne 5 : Objectif général (si défini) ────────────────────────────────
    if annee.objectif_general:
        ws.merge_cells('A5:O5')
        c = ws['A5']
        c.value = f"Objectif général : {annee.objectif_general}"
        c.font = Font(bold=True, italic=True)
        c.alignment = lft
        start_row = 6
    else:
        start_row = 5

    # En-têtes (3 niveaux) — 15 colonnes
    headers_r1 = ['Code', 'Objectifs/Résultats/Actions/Activités/Tâches', 'Imputation budgétaire',
                  'Sources de financement (F CFA)', '', '', '', '', '',
                  "Période d'exécution", 'Poids (%)',
                  'Direction/Unité Resp.', 'Structures associées / Unité Admi. Associée',
                  "Mode d'exécution", 'Observations (Contribution ODD…)']
    headers_r2 = ['', '', '',
                  'Source de financement', '', '', '', '', 'Total',
                  '', '', '', '', '', '']
    headers_r3 = ['', '', '',
                  'Fonds Propres', 'FA', 'FNA',
                  'PTFs', 'Autres Fonds', '',
                  '', '', '', '', '', '']

    hdr_fill = PatternFill("solid", fgColor="FF69B4")
    hdr_font = Font(bold=True, color="000000", size=8)

    for r_offset, row_data in enumerate([headers_r1, headers_r2, headers_r3]):
        row_num = start_row + r_offset
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=row_num, column=col, value=val)
            c.fill = hdr_fill
            c.font = hdr_font
            c.alignment = ctr
            c.border = brd

    # Fusions (15 colonnes A–O)
    ws.merge_cells(f'A{start_row}:A{start_row+2}')    # Code
    ws.merge_cells(f'B{start_row}:B{start_row+2}')    # Objectifs
    ws.merge_cells(f'C{start_row}:C{start_row+2}')    # Imputation
    ws.merge_cells(f'D{start_row}:I{start_row}')      # Sources de financement (ligne 1)
    ws.merge_cells(f'D{start_row+1}:H{start_row+1}')  # Source de financement (ligne 2)
    ws.merge_cells(f'I{start_row+1}:I{start_row+2}')  # Total (lignes 2-3)
    ws.merge_cells(f'J{start_row}:J{start_row+2}')    # Période d'exécution
    ws.merge_cells(f'K{start_row}:K{start_row+2}')    # Poids
    ws.merge_cells(f'L{start_row}:L{start_row+2}')    # Resp
    ws.merge_cells(f'M{start_row}:M{start_row+2}')    # Struct. Associées
    ws.merge_cells(f'N{start_row}:N{start_row+2}')    # Mode d'exécution
    ws.merge_cells(f'O{start_row}:O{start_row+2}')    # Observations

    row = start_row + 3
    programmes = Programme.query.filter_by(annee_id=annee.id).order_by(Programme.numero).all()
    _nb_svc = Service.query.count()
    _nb_dir = Direction.query.count()

    # Colonnes budgétaires (D=4 à I=9 inclus Total) → format numérique avec séparateur milliers
    NUM_COLS = {4, 5, 6, 7, 8, 9}

    _MOIS_COURT = {
        'Janvier': 'Janv', 'Février': 'Fév', 'Mars': 'Mars', 'Avril': 'Avr',
        'Mai': 'Mai', 'Juin': 'Juin', 'Juillet': 'Juil', 'Août': 'Août',
        'Septembre': 'Sept', 'Octobre': 'Oct', 'Novembre': 'Nov', 'Décembre': 'Déc',
    }

    def fmt_periode(debut, fin):
        d = _MOIS_COURT.get(debut or '', debut or '')
        f = _MOIS_COURT.get(fin or '', fin or '')
        if d and f and d != f:
            return f"{d} - {f}"
        return d or f or ''

    def write_row(vals, fill, bold, row_num):
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row_num, column=col, value=v)
            c.fill = fill
            c.font = Font(bold=bold, size=9)
            c.alignment = lft if col == 2 else ctr
            c.border = brd
            if col in NUM_COLS and isinstance(v, (int, float)):
                c.number_format = '#,##0'

    def fmt(v): return v if v else 0

    for prog in programmes:
        if prog.objectif_specifique:
            ws.merge_cells(f'A{row}:O{row}')
            c = ws.cell(row=row, column=1, value=f"Objectif {prog.numero}/Résultat {prog.numero} : {prog.objectif_specifique}")
            c.font = Font(bold=True, italic=True, size=9)
            c.fill = PatternFill("solid", fgColor="EAF4FB")
            c.alignment = lft
            row += 1

        write_row([prog.code, prog.libelle, '',
                   fmt(prog.ressources_propres), fmt(prog.fadec_affecte),
                   fmt(prog.fadec_non_affecte), fmt(prog.autres_partenaires),
                   fmt(prog.autres_fonds), fmt(prog.budget_total),
                   '', prog.poids, '', '', '', ''],
                  f_prog, True, row)
        row += 1

        for projet in prog.projets:
            write_row([projet.code, projet.libelle, '',
                       fmt(projet.ressources_propres), fmt(projet.fadec_affecte),
                       fmt(projet.fadec_non_affecte), fmt(projet.autres_partenaires),
                       fmt(projet.autres_fonds), fmt(projet.budget_total),
                       '', projet.poids, '', '', '', ''],
                      f_proj, True, row)
            row += 1

            for act in projet.activites:
                # Filtre par nature
                _act_inv = act.type_activite == "Activité d'investissement"
                if nature_filtre == 'inv' and not _act_inv:
                    continue
                if nature_filtre == 'fct' and _act_inv:
                    continue
                resp = act.direction_responsable.code if act.direction_responsable else ''
                nb_sa = len(act.services_intervenants)
                nb_da = len(act.directions_associees)
                all_sa = nb_sa == _nb_svc and nb_sa > 0
                all_da = nb_da == _nb_dir and nb_da > 0
                if all_sa and all_da:
                    assoc = 'Tous services et directions'
                elif all_sa:
                    assoc = 'Tous les services'
                elif all_da:
                    assoc = 'Toutes les directions'
                elif nb_sa == 0 and nb_da == 0:
                    assoc_parts = []
                    for se in act.structures_externes:
                        assoc_parts.append(se.nom)
                    if act.acteurs_externes:
                        assoc_parts.append(act.acteurs_externes)
                    assoc = ', '.join(assoc_parts) if assoc_parts else 'Aucun'
                else:
                    assoc_parts = ([s.code for s in act.services_intervenants] +
                                   [d.code for d in act.directions_associees] +
                                   [se.nom for se in act.structures_externes])
                    if act.acteurs_externes:
                        assoc_parts.append(act.acteurs_externes)
                    assoc = ', '.join(assoc_parts)
                t_rp = sum(t.ressources_propres or 0 for t in act.taches) if act.taches else act.ressources_propres or 0
                t_fa = sum(t.fadec_affecte or 0 for t in act.taches) if act.taches else act.fadec_affecte or 0
                t_fn = sum(t.fadec_non_affecte or 0 for t in act.taches) if act.taches else act.fadec_non_affecte or 0
                t_ap = sum(t.autres_partenaires or 0 for t in act.taches) if act.taches else act.autres_partenaires or 0
                t_af = sum(t.autres_fonds or 0 for t in act.taches) if act.taches else act.autres_fonds or 0
                write_row([act.code, act.nom, act.imputation_budgetaire or 'NÉANT',
                           t_rp, t_fa, t_fn, t_ap, t_af, t_rp+t_fa+t_fn+t_ap+t_af,
                           fmt_periode(act.periode_debut, act.periode_fin),
                           act.poids, resp, assoc, act.mode_execution or 'Direct',
                           act.observations or ''],
                          f_act, True, row)
                row += 1

                for tache in act.taches:
                    if service_filtre and tache.service_responsable_id != service_filtre:
                        continue
                    if tache.service_responsable:
                        resp_t = tache.service_responsable.code
                    elif tache.direction_responsable:
                        resp_t = tache.direction_responsable.code
                    else:
                        resp_t = ''
                    nb_st = len(tache.services_concernes)
                    nb_dt = len(tache.directions_associees)
                    all_st = nb_st == _nb_svc and nb_st > 0
                    all_dt = nb_dt == _nb_dir and nb_dt > 0
                    if all_st and all_dt:
                        assoc_t = 'Tous services et directions'
                    elif all_st:
                        assoc_t = 'Tous les services'
                    elif all_dt:
                        assoc_t = 'Toutes les directions'
                    elif nb_st == 0 and nb_dt == 0:
                        assoc_t_parts = [se.nom for se in tache.structures_externes]
                        if tache.acteurs_externes:
                            assoc_t_parts.append(tache.acteurs_externes)
                        assoc_t = ', '.join(assoc_t_parts) if assoc_t_parts else 'Aucun'
                    else:
                        assoc_t_parts = ([s.code for s in tache.services_concernes] +
                                         [d.code for d in tache.directions_associees] +
                                         [se.nom for se in tache.structures_externes])
                        if tache.acteurs_externes:
                            assoc_t_parts.append(tache.acteurs_externes)
                        assoc_t = ', '.join(assoc_t_parts)
                    write_row([tache.code, f"  {tache.nom}",
                               tache.imputation_budgetaire or 'NÉANT',
                               tache.ressources_propres or 0, tache.fadec_affecte or 0,
                               tache.fadec_non_affecte or 0, tache.autres_partenaires or 0,
                               tache.autres_fonds or 0, tache.budget_total,
                               fmt_periode(tache.periode_debut, tache.periode_fin),
                               tache.poids, resp_t, assoc_t,
                               tache.mode_execution or 'Direct',
                               tache.observations or ''],
                              f_tch, False, row)
                    row += 1

    # Total général (réutilise la liste déjà chargée)
    tg_rp = sum(p.ressources_propres for p in programmes)
    tg_fa = sum(p.fadec_affecte for p in programmes)
    tg_fn = sum(p.fadec_non_affecte for p in programmes)
    tg_ap = sum(p.autres_partenaires for p in programmes)
    tg_af = sum(p.autres_fonds for p in programmes)
    write_row(['', 'TOTAL GÉNÉRAL', '',
               tg_rp, tg_fa, tg_fn, tg_ap, tg_af, tg_rp+tg_fa+tg_fn+tg_ap+tg_af,
               '', 100, '', '', '', ''],
              PatternFill("solid", fgColor="AED6F1"),
              True, row)
    for col in range(1, 16):
        ws.cell(row=row, column=col).font = Font(bold=True, color="000000", size=9)

    # ── Ligne de pied de tableau ───────────────────────────────────────────────
    row += 1
    from datetime import datetime as _dt
    _date_str = _dt.now().strftime('%d/%m/%Y à %H:%M')
    ws.merge_cells(f'A{row}:G{row}')
    c = ws.cell(row=row, column=1, value=f"Exporté le {_date_str}")
    c.font = Font(bold=True, size=8)
    c.alignment = lft
    ws.merge_cells(f'H{row}:O{row}')
    c = ws.cell(row=row, column=8,
                value="Direction du Développement Local et de la Planification (DDLP)")
    c.font = Font(bold=True, size=8)
    c.alignment = Alignment(horizontal='right', vertical='center')

    # Largeurs colonnes (15 colonnes A–O)
    widths = [10, 45, 14, 13, 10, 10, 13, 13, 14, 14, 8, 18, 28, 18, 28]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = f'A{start_row + 3}'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    fname = f"PTA_{annee.annee}.xlsx"
    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=fname)


# ─── Réinitialisation PTA ────────────────────────────────────────────────────

def _creer_backup_pta(annee):
    """Sérialise tout le PTA de l'année en JSON et l'enregistre en base."""
    def ser_tache(t):
        return {
            'numero': t.numero, 'ordre': t.ordre or 0, 'nom': t.nom,
            'description': t.description, 'poids': t.poids or 0,
            'imputation_budgetaire': t.imputation_budgetaire,
            'ressources_propres': t.ressources_propres or 0,
            'fadec_affecte': t.fadec_affecte or 0,
            'fadec_non_affecte': t.fadec_non_affecte or 0,
            'autres_partenaires': t.autres_partenaires or 0,
            'autres_fonds': t.autres_fonds or 0,
            'details_financement': t.details_financement,
            'acteurs_externes': t.acteurs_externes,
            'mode_execution': t.mode_execution,
            'service_responsable_id': t.service_responsable_id,
            'direction_responsable_id': t.direction_responsable_id,
            'services_concernes': [s.id for s in t.services_concernes],
            'directions_associees': [d.id for d in t.directions_associees],
            'structures_externes': [se.id for se in t.structures_externes],
        }

    def ser_activite(a):
        return {
            'numero': a.numero, 'nom': a.nom, 'description': a.description,
            'direction_responsable_id': a.direction_responsable_id,
            'imputation_budgetaire': a.imputation_budgetaire,
            'ressources_propres': a.ressources_propres or 0,
            'fadec_affecte': a.fadec_affecte or 0,
            'fadec_non_affecte': a.fadec_non_affecte or 0,
            'autres_partenaires': a.autres_partenaires or 0,
            'autres_fonds': a.autres_fonds or 0,
            'details_financement': a.details_financement,
            'acteurs_externes': a.acteurs_externes,
            'periode_debut': a.periode_debut, 'periode_fin': a.periode_fin,
            'mode_execution': a.mode_execution, 'poids': a.poids or 0,
            'services_intervenants': [s.id for s in a.services_intervenants],
            'directions_associees': [d.id for d in a.directions_associees],
            'structures_externes': [se.id for se in a.structures_externes],
            'taches': [ser_tache(t) for t in a.taches],
        }

    def ser_projet(pj):
        return {
            'numero': pj.numero, 'nom': pj.nom,
            'description': pj.description, 'poids': pj.poids or 0,
            'activites': [ser_activite(a) for a in pj.activites],
        }

    def ser_programme(p):
        return {
            'numero': p.numero, 'nom': p.nom, 'description': p.description,
            'objectif_specifique': p.objectif_specifique, 'poids': p.poids or 0,
            'projets': [ser_projet(pj) for pj in p.projets],
        }

    from datetime import datetime
    programmes = Programme.query.filter_by(annee_id=annee.id).order_by(Programme.numero).all()
    nb_acts = sum(len(pj.activites) for p in programmes for pj in p.projets)
    nb_taches = sum(len(a.taches) for p in programmes for pj in p.projets for a in pj.activites)

    data = {
        'annee': {'id': annee.id, 'annee': annee.annee,
                  'objectif_general': annee.objectif_general},
        'programmes': [ser_programme(p) for p in programmes],
        'backup_at': datetime.utcnow().isoformat(),
    }

    backup = PTABackup(
        annee_id=annee.id,
        annee_label=annee.annee,
        created_by_id=current_user.id,
        nb_programmes=len(programmes),
        nb_activites=nb_acts,
        nb_taches=nb_taches,
        contenu=json.dumps(data, ensure_ascii=False),
    )
    db.session.add(backup)
    return backup


@pta_bp.route('/save_manual', methods=['POST'])
@editeur_only
def pta_save_manual():
    annee = get_annee()
    if not annee:
        flash('Aucune année active.', 'danger')
        return redirect(url_for('pta.global_pta'))
    backup = _creer_backup_pta(annee)
    db.session.commit()
    flash(
        f'Sauvegarde du PTA {annee.annee} créée (backup #{backup.id} — '
        f'{backup.nb_programmes} programme(s), {backup.nb_activites} activité(s), '
        f'{backup.nb_taches} tâche(s)). Récupérable depuis Administration › Sauvegardes.',
        'success'
    )
    return redirect(request.referrer or url_for('pta.global_pta'))


@pta_bp.route('/reset', methods=['GET', 'POST'])
@editeur_only
def pta_reset():
    annee = get_annee()
    if not annee:
        flash('Aucune année active.', 'danger')
        return redirect(url_for('pta.global_pta'))

    programmes = Programme.query.filter_by(annee_id=annee.id).order_by(Programme.numero).all()
    nb_progs = len(programmes)
    nb_acts = sum(len(pj.activites) for p in programmes for pj in p.projets)
    nb_taches = sum(len(a.taches) for p in programmes for pj in p.projets for a in pj.activites)

    if request.method == 'POST':
        password = request.form.get('password', '').strip()
        if not current_user.check_password(password):
            flash('Mot de passe incorrect.', 'danger')
            return redirect(url_for('pta.pta_reset'))

        backup = _creer_backup_pta(annee)
        for p in programmes:
            db.session.delete(p)
        db.session.commit()
        flash(
            f'PTA {annee.annee} réinitialisé ({nb_progs} programmes, {nb_acts} activités, '
            f'{nb_taches} tâches supprimés). Backup #{backup.id} créé — '
            f'récupérable depuis Administration › Sauvegardes.',
            'success'
        )
        return redirect(url_for('pta.global_pta'))

    return render_template('pta/reset_confirm.html',
                           annee=annee,
                           nb_progs=nb_progs,
                           nb_acts=nb_acts,
                           nb_taches=nb_taches)