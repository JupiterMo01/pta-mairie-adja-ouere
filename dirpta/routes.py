import io
from flask import render_template, redirect, url_for, flash, request, session, send_file
from flask_login import login_required, current_user
from models import db, Programme, Direction, Annee, Service
from dirpta import dirpta_bp


def get_annee():
    annee_id = session.get('annee_id')
    if annee_id:
        return Annee.query.get(annee_id)
    return Annee.query.filter_by(actif=True).first()


# ── Renormalisation des poids ─────────────────────────────────────────────────

def _renorm(items, src='original_poids', dst='new_poids'):
    """Recalcule les poids pour que leur somme soit exactement 100.
    Le dernier élément absorbe les micro-erreurs d'arrondi."""
    total = sum(i[src] for i in items)
    if not items:
        return
    if total == 0:
        eq = round(100 / len(items), 4)
        for i in items:
            i[dst] = eq
        return
    running = 0.0
    for idx, i in enumerate(items):
        if idx == len(items) - 1:
            i[dst] = round(100.0 - running, 4)
        else:
            p = round(i[src] / total * 100, 4)
            i[dst] = p
            running += p


# ── Calcul du PTA filtré pour une direction ───────────────────────────────────

def _compute_pta_direction(annee, direction):
    """
    Retourne la hiérarchie Programme→Projet→Activité→Tâche filtrée pour
    la direction donnée (tâches dont direction_responsable_id == direction.id).
    Les poids sont renormalisés à chaque niveau pour toujours sommer à 100.
    Les codes sont renumérotés indépendamment du PTA global.
    """
    programmes = Programme.query.filter_by(annee_id=annee.id)\
                                .order_by(Programme.numero).all()
    result_progs = []

    for prog in programmes:
        result_projs = []

        for proj in prog.projets:
            result_acts = []

            for act in proj.activites:
                dir_taches = [
                    t for t in sorted(act.taches, key=lambda x: (x.ordre or 0, x.id))
                    if t.direction_responsable_id == direction.id
                ]
                if not dir_taches:
                    continue

                taches_data = [{'tache': t, 'original_poids': t.poids or 0}
                               for t in dir_taches]
                _renorm(taches_data)

                result_acts.append({
                    'activite': act,
                    'taches':   taches_data,
                    'original_poids': act.poids or 0,
                })

            if not result_acts:
                continue
            _renorm(result_acts)

            result_projs.append({
                'projet':   proj,
                'activites': result_acts,
                'original_poids': proj.poids or 0,
            })

        if not result_projs:
            continue
        _renorm(result_projs)

        result_progs.append({
            'programme': prog,
            'projets':   result_projs,
            'original_poids': prog.poids or 0,
        })

    if result_progs:
        _renorm(result_progs)

    # ── Renumérotation ────────────────────────────────────────────────────────
    for pi, pd in enumerate(result_progs, 1):
        pd['code'] = str(pi)
        for pji, pjd in enumerate(pd['projets'], 1):
            pjd['code'] = f"{pi}.{pji}"
            for ai, ad in enumerate(pjd['activites'], 1):
                ad['code'] = f"{pi}.{pji}.{ai}"
                for ti, td in enumerate(ad['taches'], 1):
                    td['num'] = f"{pi}.{pji}.{ai}.{ti}"

    return result_progs


# ── Export Excel ──────────────────────────────────────────────────────────────

def _fill_sheet(ws, annee, direction, data):
    """Remplit une feuille Excel avec le PTA d'une direction.
    Peut être appelée pour une feuille isolée ou dans un classeur multi-feuilles."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime as _dt
    import os as _os
    from flask import current_app

    thin = Side(style='thin')
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    lft  = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    f_prog = PatternFill("solid", fgColor="FFFF99")
    f_proj = PatternFill("solid", fgColor="FFB6C1")
    f_act  = PatternFill("solid", fgColor="D3D3D3")
    f_tch  = PatternFill("solid", fgColor="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="FF69B4")
    os_fill  = PatternFill("solid", fgColor="FFFACD")
    tot_fill = PatternFill("solid", fgColor="AED6F1")

    NUM_COLS = {4, 5, 6, 7, 8, 9}

    _MOIS_COURT = {
        'Janvier': 'Janv', 'Février': 'Fév', 'Mars': 'Mars', 'Avril': 'Avr',
        'Mai': 'Mai', 'Juin': 'Juin', 'Juillet': 'Juil', 'Août': 'Août',
        'Septembre': 'Sept', 'Octobre': 'Oct', 'Novembre': 'Nov', 'Décembre': 'Déc',
    }

    def fmt_periode(debut, fin):
        d = _MOIS_COURT.get(debut or '', debut or '')
        f = _MOIS_COURT.get(fin   or '', fin   or '')
        if d and f and d != f:
            return f"{d} - {f}"
        return d or f or ''

    def fmt(v):
        return v if v else 0

    def write_row(vals, fill, bold, row_num):
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row_num, column=col, value=v)
            c.fill = fill
            c.font = Font(bold=bold, size=9)
            c.alignment = lft if col == 2 else ctr
            c.border = brd
            if col in NUM_COLS and isinstance(v, (int, float)):
                c.number_format = '#,##0'

    # ── Lignes 1-3 : En-tête institutionnel (bandeau + contacts + logo) ────────
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 8
    ws.row_dimensions[3].height = 22

    ws.merge_cells('A1:D3')
    ws.merge_cells('E1:J3')
    c = ws['E1']
    c.value = "BP 02 Adja-Ouèrè\nTél : +229 01 61 91 96 12\nEmail : contact.adjaouere@mairie.bj"
    c.font = Font(bold=True, size=9)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.merge_cells('K1:O3')

    try:
        from openpyxl.drawing.image import Image as _XLImg
        from PIL import Image as _PILImg
        import io as _imgio
        _static_img = _os.path.join(current_app.root_path, 'static', 'img')
        _H = 52
        _band_path = _os.path.join(_static_img, 'bandeau.png')
        _logo_path = _os.path.join(_static_img, 'logo_commune.png')
        if _os.path.exists(_band_path):
            _buf = _imgio.BytesIO()
            with _PILImg.open(_band_path) as _pil:
                _ow, _oh = _pil.size
                _pil.save(_buf, format='PNG')
            _buf.seek(0)
            _i = _XLImg(_buf)
            _i.height = _H
            _i.width = int(_ow * _H / _oh)
            ws.add_image(_i, 'A1')
        if _os.path.exists(_logo_path):
            _buf2 = _imgio.BytesIO()
            with _PILImg.open(_logo_path) as _pil2:
                _ow2, _oh2 = _pil2.size
                _pil2.save(_buf2, format='PNG')
            _buf2.seek(0)
            _i2 = _XLImg(_buf2)
            _lw = int(_ow2 * _H / _oh2)
            _i2.height = _H
            _i2.width = _lw
            try:
                from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
                from openpyxl.drawing.xdr import XDRPositiveSize2D
                _anch2 = OneCellAnchor()
                _anch2._from = AnchorMarker(col=15, colOff=-int(_lw * 9525), row=0, rowOff=0)
                _anch2.ext = XDRPositiveSize2D(int(_lw * 9525), int(_H * 9525))
                _i2.anchor = _anch2
                ws.add_image(_i2)
            except Exception:
                ws.add_image(_i2, 'N1')
    except Exception:
        pass

    # ── Ligne 4 : Titre PTA + Direction ────────────────────────────────────────
    ws.merge_cells('A4:O4')
    ws['A4'].value = (f"PLAN DE TRAVAIL ANNUEL — Exercice {annee.annee}"
                      f"   |   Direction : {direction.code} — {direction.nom}")
    ws['A4'].font = Font(bold=True, size=13)
    ws['A4'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A4'].fill = PatternFill("solid", fgColor="D6EAF8")
    _med = Side(style='medium')
    for _tc in range(1, 16):
        _cell2 = ws.cell(row=4, column=_tc)
        _cell2.border = Border(
            left=_med  if _tc == 1  else Side(style=None),
            right=_med if _tc == 15 else Side(style=None),
            top=_med, bottom=_med
        )
    ws.row_dimensions[4].height = 28

    # ── Ligne 5 : Objectif général ─────────────────────────────────────────────
    if annee.objectif_general:
        ws.merge_cells('A5:O5')
        c = ws['A5']
        c.value = f"Objectif général : {annee.objectif_general}"
        c.font = Font(bold=True, italic=True)
        c.alignment = lft
        start_row = 6
    else:
        start_row = 5

    # ── En-têtes du tableau (3 lignes, 15 colonnes A–O) ────────────────────────
    headers_r1 = ['Code', 'Objectifs/Résultats/Actions/Activités/Tâches', 'Imputation budgétaire',
                  'Sources de financement (F CFA)', '', '', '', '', '',
                  "Période d'exécution", 'Poids (%)',
                  'Direction/Unité Resp.', 'Structures associées / Unité Admi. Associée',
                  "Mode d'exécution", 'Observations (Contribution ODD…)']
    headers_r2 = ['', '', '', 'Source de financement', '', '', '', '', 'Total', '', '', '', '', '', '']
    headers_r3 = ['', '', '', 'Fonds Propres', 'FA', 'FNA', 'PTFs', 'Autres Fonds', '', '', '', '', '', '', '']

    for r_offset, row_data in enumerate([headers_r1, headers_r2, headers_r3]):
        rn = start_row + r_offset
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=rn, column=col, value=val)
            c.fill = hdr_fill
            c.font = Font(bold=True, color="000000", size=8)
            c.alignment = ctr
            c.border = brd

    ws.merge_cells(f'A{start_row}:A{start_row+2}')
    ws.merge_cells(f'B{start_row}:B{start_row+2}')
    ws.merge_cells(f'C{start_row}:C{start_row+2}')
    ws.merge_cells(f'D{start_row}:I{start_row}')
    ws.merge_cells(f'D{start_row+1}:H{start_row+1}')
    ws.merge_cells(f'I{start_row+1}:I{start_row+2}')
    ws.merge_cells(f'J{start_row}:J{start_row+2}')
    ws.merge_cells(f'K{start_row}:K{start_row+2}')
    ws.merge_cells(f'L{start_row}:L{start_row+2}')
    ws.merge_cells(f'M{start_row}:M{start_row+2}')
    ws.merge_cells(f'N{start_row}:N{start_row+2}')
    ws.merge_cells(f'O{start_row}:O{start_row+2}')

    row = start_row + 3

    _nb_svc = Service.query.count()
    _nb_dir = Direction.query.count()

    def _assoc_act(act):
        nb_s = len(act.services_intervenants)
        nb_d = len(act.directions_associees)
        if nb_s == _nb_svc and nb_s > 0 and nb_d == _nb_dir and nb_d > 0:
            return 'Tous services et directions'
        if nb_s == _nb_svc and nb_s > 0:
            return 'Tous les services'
        if nb_d == _nb_dir and nb_d > 0:
            return 'Toutes les directions'
        parts = ([s.code for s in act.services_intervenants] +
                 [d.code for d in act.directions_associees] +
                 [se.nom for se in act.structures_externes])
        if act.acteurs_externes:
            parts.append(act.acteurs_externes)
        return ', '.join(parts) if parts else 'Aucun'

    def _assoc_tache(t):
        nb_s = len(t.services_concernes)
        nb_d = len(t.directions_associees)
        if nb_s == _nb_svc and nb_s > 0 and nb_d == _nb_dir and nb_d > 0:
            return 'Tous services et directions'
        if nb_s == _nb_svc and nb_s > 0:
            return 'Tous les services'
        if nb_d == _nb_dir and nb_d > 0:
            return 'Toutes les directions'
        parts = ([s.code for s in t.services_concernes] +
                 [d.code for d in t.directions_associees] +
                 [se.nom for se in t.structures_externes])
        if t.acteurs_externes:
            parts.append(t.acteurs_externes)
        return ', '.join(parts) if parts else 'Aucun'

    # ── Données ────────────────────────────────────────────────────────────────
    for pd in data:
        prog = pd['programme']

        # Objectif spécifique
        if prog.objectif_specifique:
            ws.merge_cells(f'A{row}:O{row}')
            c = ws.cell(row=row, column=1,
                        value=f"Objectif {pd['code']}/Résultat {pd['code']} : {prog.objectif_specifique}")
            c.font = Font(bold=True, italic=True, size=9)
            c.fill = os_fill
            c.alignment = lft
            row += 1

        # Totaux programme depuis tâches filtrées
        p_rp = sum(td['tache'].ressources_propres or 0 for pjd in pd['projets'] for ad in pjd['activites'] for td in ad['taches'])
        p_fa = sum(td['tache'].fadec_affecte      or 0 for pjd in pd['projets'] for ad in pjd['activites'] for td in ad['taches'])
        p_fn = sum(td['tache'].fadec_non_affecte  or 0 for pjd in pd['projets'] for ad in pjd['activites'] for td in ad['taches'])
        p_ap = sum(td['tache'].autres_partenaires or 0 for pjd in pd['projets'] for ad in pjd['activites'] for td in ad['taches'])
        p_af = sum(td['tache'].autres_fonds       or 0 for pjd in pd['projets'] for ad in pjd['activites'] for td in ad['taches'])

        write_row([pd['code'], f"Programme {pd['code']} : {prog.nom}", '',
                   fmt(p_rp), fmt(p_fa), fmt(p_fn), fmt(p_ap), fmt(p_af), fmt(p_rp+p_fa+p_fn+p_ap+p_af),
                   '', round(pd['new_poids'], 2), '', '', '', ''],
                  f_prog, True, row)
        row += 1

        for pjd in pd['projets']:
            proj = pjd['projet']

            j_rp = sum(td['tache'].ressources_propres or 0 for ad in pjd['activites'] for td in ad['taches'])
            j_fa = sum(td['tache'].fadec_affecte      or 0 for ad in pjd['activites'] for td in ad['taches'])
            j_fn = sum(td['tache'].fadec_non_affecte  or 0 for ad in pjd['activites'] for td in ad['taches'])
            j_ap = sum(td['tache'].autres_partenaires or 0 for ad in pjd['activites'] for td in ad['taches'])
            j_af = sum(td['tache'].autres_fonds       or 0 for ad in pjd['activites'] for td in ad['taches'])

            write_row([pjd['code'], f"Projet {pjd['code']} : {proj.nom}", '',
                       fmt(j_rp), fmt(j_fa), fmt(j_fn), fmt(j_ap), fmt(j_af), fmt(j_rp+j_fa+j_fn+j_ap+j_af),
                       '', round(pjd['new_poids'], 2), '', '', '', ''],
                      f_proj, True, row)
            row += 1

            for ad in pjd['activites']:
                act = ad['activite']
                a_rp = sum(td['tache'].ressources_propres or 0 for td in ad['taches'])
                a_fa = sum(td['tache'].fadec_affecte      or 0 for td in ad['taches'])
                a_fn = sum(td['tache'].fadec_non_affecte  or 0 for td in ad['taches'])
                a_ap = sum(td['tache'].autres_partenaires or 0 for td in ad['taches'])
                a_af = sum(td['tache'].autres_fonds       or 0 for td in ad['taches'])

                write_row([ad['code'], act.nom, act.imputation_budgetaire or 'NÉANT',
                           fmt(a_rp), fmt(a_fa), fmt(a_fn), fmt(a_ap), fmt(a_af),
                           fmt(a_rp+a_fa+a_fn+a_ap+a_af),
                           fmt_periode(act.periode_debut, act.periode_fin),
                           round(ad['new_poids'], 2),
                           direction.code,
                           _assoc_act(act),
                           act.mode_execution or 'Direct',
                           act.observations or ''],
                          f_act, True, row)
                row += 1

                for td in ad['taches']:
                    t = td['tache']
                    resp_t = t.service_responsable.code if t.service_responsable else direction.code
                    write_row([td['num'], f"  {t.nom}",
                               t.imputation_budgetaire or 'NÉANT',
                               fmt(t.ressources_propres), fmt(t.fadec_affecte),
                               fmt(t.fadec_non_affecte), fmt(t.autres_partenaires),
                               fmt(t.autres_fonds), fmt(t.budget_total),
                               fmt_periode(t.periode_debut, t.periode_fin),
                               round(td['new_poids'], 2),
                               resp_t,
                               _assoc_tache(t),
                               t.mode_execution or 'Direct',
                               t.observations or ''],
                              f_tch, False, row)
                    row += 1

    # ── Total général ──────────────────────────────────────────────────────────
    tg_rp = sum(td['tache'].ressources_propres or 0 for pd in data for pjd in pd['projets'] for ad in pjd['activites'] for td in ad['taches'])
    tg_fa = sum(td['tache'].fadec_affecte      or 0 for pd in data for pjd in pd['projets'] for ad in pjd['activites'] for td in ad['taches'])
    tg_fn = sum(td['tache'].fadec_non_affecte  or 0 for pd in data for pjd in pd['projets'] for ad in pjd['activites'] for td in ad['taches'])
    tg_ap = sum(td['tache'].autres_partenaires or 0 for pd in data for pjd in pd['projets'] for ad in pjd['activites'] for td in ad['taches'])
    tg_af = sum(td['tache'].autres_fonds       or 0 for pd in data for pjd in pd['projets'] for ad in pjd['activites'] for td in ad['taches'])
    write_row(['', 'TOTAL GÉNÉRAL', '',
               tg_rp, tg_fa, tg_fn, tg_ap, tg_af, tg_rp+tg_fa+tg_fn+tg_ap+tg_af,
               '', 100, '', '', '', ''],
              tot_fill, True, row)
    for col in range(1, 16):
        ws.cell(row=row, column=col).font = Font(bold=True, color="000000", size=9)

    # ── Pied de tableau ────────────────────────────────────────────────────────
    row += 1
    _date_str = _dt.now().strftime('%d/%m/%Y à %H:%M')
    ws.merge_cells(f'A{row}:G{row}')
    c = ws.cell(row=row, column=1, value=f"Exporté le {_date_str}")
    c.font = Font(bold=True, size=8)
    c.alignment = lft
    ws.merge_cells(f'H{row}:O{row}')
    c = ws.cell(row=row, column=8,
                value="Direction du Développement Local et de la Planification (DDLP)")
    c.font = Font(bold=True, size=8)
    c.alignment = Alignment(horizontal='right', vertical='center')

    # ── Largeurs colonnes (15 colonnes A–O, identiques au PTA global) ─────────
    widths = [10, 45, 14, 13, 10, 10, 13, 13, 14, 14, 8, 18, 28, 18, 28]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = f'A{start_row + 3}'


def _build_excel(annee, direction, data):
    """Crée un classeur Excel avec une seule feuille pour la direction."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = (direction.code[:31])
    _fill_sheet(ws, annee, direction, data)
    return wb


# ── Routes ────────────────────────────────────────────────────────────────────

@dirpta_bp.route('/', methods=['GET'])
@login_required
def index():
    annee = get_annee()
    is_direction_user = (current_user.role == 'direction')

    if is_direction_user:
        # L'agent de direction ne voit que SA direction — pas de sélecteur
        direction  = current_user.direction
        directions = []
    else:
        directions = Direction.query.order_by(Direction.nom).all()
        dir_id     = request.args.get('direction_id', type=int)
        direction  = Direction.query.get(dir_id) if dir_id else None

    # Services rattachés à la direction (pour filtre direction user)
    dir_services = sorted(direction.services, key=lambda s: s.nom) if direction else []

    # Filtre optionnel par service (direction user uniquement)
    sel_svc_id  = request.args.get('service_id', type=int) if is_direction_user else None
    sel_service = None
    if sel_svc_id:
        _svc = Service.query.get(sel_svc_id)
        if _svc and direction and _svc.direction_id == direction.id:
            sel_service = _svc

    data = []
    if annee and direction:
        if sel_service:
            from svcpta.routes import _compute_pta_service
            data = _compute_pta_service(annee, sel_service)
        else:
            data = _compute_pta_direction(annee, direction)

    return render_template('dirpta/index.html',
                           annee=annee, directions=directions,
                           direction=direction, data=data,
                           is_direction_user=is_direction_user,
                           dir_services=dir_services,
                           sel_service=sel_service)


@dirpta_bp.route('/excel-toutes-directions')
@login_required
def export_excel_all():
    """Exporte le PTA de toutes les directions dans un seul classeur (1 feuille par direction)."""
    from openpyxl import Workbook
    annee = get_annee()
    if not annee:
        flash('Aucune année active.', 'danger')
        return redirect(url_for('dirpta.index'))

    directions = Direction.query.order_by(Direction.nom).all()

    wb = Workbook()
    wb.remove(wb.active)   # supprimer la feuille vide créée par défaut

    nb_feuilles = 0
    for d in directions:
        data = _compute_pta_direction(annee, d)
        if not data:
            continue
        ws = wb.create_sheet(title=d.code[:31])
        _fill_sheet(ws, annee, d, data)
        nb_feuilles += 1

    if nb_feuilles == 0:
        flash('Aucune direction n\'a de tâches dans le PTA.', 'warning')
        return redirect(url_for('dirpta.index'))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    fname = f"PTA_Toutes_Directions_{annee.annee}.xlsx"
    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=fname)


@dirpta_bp.route('/<int:direction_id>/excel')
@login_required
def export_excel(direction_id):
    annee     = get_annee()
    direction = Direction.query.get_or_404(direction_id)
    if not annee:
        flash('Aucune année active.', 'danger')
        return redirect(url_for('dirpta.index'))

    data = _compute_pta_direction(annee, direction)
    wb   = _build_excel(annee, direction, data)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    fname = f"PTA_{direction.code}_{annee.annee}.xlsx"
    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=fname)


@dirpta_bp.route('/<int:direction_id>/print')
@login_required
def print_view(direction_id):
    annee     = get_annee()
    direction = Direction.query.get_or_404(direction_id)
    if not annee:
        flash('Aucune année active.', 'danger')
        return redirect(url_for('dirpta.index'))
    data      = _compute_pta_direction(annee, direction)
    services   = Service.query.order_by(Service.nom).all()
    directions = Direction.query.order_by(Direction.nom).all()
    return render_template('dirpta/print_view.html',
                           annee=annee, direction=direction, data=data,
                           services=services, directions=directions)
