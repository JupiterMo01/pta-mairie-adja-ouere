"""
suivi/routes.py

Vues et exports du module Suivi & Évaluation.

Niveaux d'accès :
  admin     → Global (tout) | par direction | par service   — lecture seule
  direction → Total direction (direct + services) | filtre par service de la direction (lecture)
              → peut modifier toutes ses tâches (filtrées par direction_responsable_id)
  service   → Ses propres tâches uniquement                 — tout modifiable

Onglets : Global | T1 | T2 | T3 | T4  (filtrage par période d'exécution des tâches)
"""
import io
from datetime import datetime
from flask import render_template, request, session, jsonify, redirect, url_for, flash, send_file
from flask_login import login_required, current_user

from models import db, Programme, Tache, SuiviTache, Annee, Direction, Service
from svcpta.routes import _compute_pta_service, _renorm
from dirpta.routes import _compute_pta_direction
from suivi import suivi_bp
from utils import log_audit, get_annee, MOIS_COURT, MOIS_ORDRE, TRIMESTRE_RANGE


# ─────────────────────────────────────────────────────────────────────────────
#  Helpers PTA
# ─────────────────────────────────────────────────────────────────────────────

def _get_annee():
    return get_annee()


def _compute_pta_global(annee):
    """
    Retourne toutes les tâches du PTA (aucun filtre service/direction).
    Même structure de données que _compute_pta_service / _compute_pta_direction.
    """
    programmes = Programme.query.filter_by(annee_id=annee.id)\
                                .order_by(Programme.numero).all()
    result_progs = []

    for prog in programmes:
        result_projs = []
        for proj in sorted(prog.projets, key=lambda p: p.numero):
            result_acts = []
            for act in sorted(proj.activites, key=lambda a: a.numero):
                taches = sorted(act.taches, key=lambda t: (t.ordre or 0, t.id))
                if not taches:
                    continue
                taches_data = [{'tache': t, 'original_poids': t.poids or 0} for t in taches]
                _renorm(taches_data)
                result_acts.append({'activite': act, 'taches': taches_data,
                                    'original_poids': act.poids or 0})
            if not result_acts:
                continue
            _renorm(result_acts)
            result_projs.append({'projet': proj, 'activites': result_acts,
                                 'original_poids': proj.poids or 0})
        if not result_projs:
            continue
        _renorm(result_projs)
        result_progs.append({'programme': prog, 'projets': result_projs,
                             'original_poids': prog.poids or 0})

    if result_progs:
        _renorm(result_progs)

    for pi, pd in enumerate(result_progs, 1):
        pd['code'] = str(pi)
        for pji, pjd in enumerate(pd['projets'], 1):
            pjd['code'] = f"{pi}.{pji}"
            for ai, ad in enumerate(pjd['activites'], 1):
                ad['code'] = f"{pi}.{pji}.{ai}"
                for ti, td in enumerate(ad['taches'], 1):
                    td['num'] = f"{pi}.{pji}.{ai}.{ti}"

    return result_progs


# ─────────────────────────────────────────────────────────────────────────────
#  Helpers suivi
# ─────────────────────────────────────────────────────────────────────────────

def _task_in_trimestre(tache, trimestre):
    debut = MOIS_ORDRE.get(tache.periode_debut or '', 0)
    fin   = MOIS_ORDRE.get(tache.periode_fin   or '', 0)
    if debut == 0 and fin == 0:
        return True
    if debut == 0: debut = fin
    if fin   == 0: fin   = debut
    t0, t1 = TRIMESTRE_RANGE[trimestre]
    return debut <= t1 and fin >= t0


def _filter_and_renorm(data, trimestre):
    """Filtre par période et renormalise. trimestre=0 → tout (pas de filtre)."""
    if trimestre == 0:
        return data

    result_progs = []
    for pd in data:
        result_projs = []
        for pjd in pd['projets']:
            result_acts = []
            for ad in pjd['activites']:
                kept = [td for td in ad['taches']
                        if _task_in_trimestre(td['tache'], trimestre)]
                if not kept:
                    continue
                _renorm(kept, src='original_poids', dst='new_poids')
                result_acts.append(dict(ad, taches=kept))
            if not result_acts:
                continue
            _renorm(result_acts, src='original_poids', dst='new_poids')
            result_projs.append(dict(pjd, activites=result_acts))
        if not result_projs:
            continue
        _renorm(result_projs, src='original_poids', dst='new_poids')
        result_progs.append(dict(pd, projets=result_projs))

    if result_progs:
        _renorm(result_progs, src='original_poids', dst='new_poids')

    # Renumérotation après filtrage
    for pi, pd2 in enumerate(result_progs, 1):
        pd2['code'] = str(pi)
        for pji, pjd2 in enumerate(pd2['projets'], 1):
            pjd2['code'] = f"{pi}.{pji}"
            for ai, ad2 in enumerate(pjd2['activites'], 1):
                ad2['code'] = f"{pi}.{pji}.{ai}"
                for ti, td2 in enumerate(ad2['taches'], 1):
                    td2['num'] = f"{pi}.{pji}.{ai}.{ti}"

    return result_progs


def _taux_suivi(sv):
    if sv is None:              return 0.0
    if sv.statut == 'execute':  return 100.0
    if sv.statut == 'en_cours': return float(sv.taux_execution or 0)
    return 0.0


def _statut_agrege(statuts):
    if not statuts:                          return 'non_execute'
    if all(s == 'execute'     for s in statuts): return 'execute'
    if all(s == 'non_execute' for s in statuts): return 'non_execute'
    return 'en_cours'


def _load_suivis_trimestre(trimestre, annee_id):
    suivis = SuiviTache.query.filter_by(trimestre=trimestre, annee_id=annee_id).all()
    result = {}
    for s in suivis:
        ex = result.get(s.tache_id)
        if ex is None or (s.service_id is not None and ex.service_id is None):
            result[s.tache_id] = s
    return result


def _load_suivis_global(annee_id):
    """Garde le suivi du trimestre le plus récent pour chaque tâche."""
    suivis = SuiviTache.query.filter_by(annee_id=annee_id)\
                             .order_by(SuiviTache.trimestre.desc()).all()
    result = {}
    for s in suivis:
        ex = result.get(s.tache_id)
        if ex is None:
            result[s.tache_id] = s
        elif s.trimestre == ex.trimestre and s.service_id is not None and ex.service_id is None:
            result[s.tache_id] = s
    return result


def _enrich(data, suivi_map, peut_modifier_fn=None):
    """
    Greffe suivi/taux/statut sur chaque tâche. Calcule les taux agrégés.
    peut_modifier_fn : callable(tache)->bool | True (tout éditable) | None (tout lecture seule)
    Retourne le taux global (float) et un bool indiquant s'il y a au moins une tâche éditable.
    """
    global_items   = []
    has_editable   = False

    for pd in data:
        prog_items = []
        for pjd in pd['projets']:
            proj_items = []
            for ad in pjd['activites']:
                act_items   = []
                act_statuts = []
                for td in ad['taches']:
                    sv = suivi_map.get(td['tache'].id)
                    tx = _taux_suivi(sv)
                    st = sv.statut if sv else 'non_execute'
                    td['suivi']  = sv
                    td['taux']   = tx
                    td['statut'] = st

                    if peut_modifier_fn is None:
                        td['peut_modifier'] = False
                    elif peut_modifier_fn is True:
                        td['peut_modifier'] = True
                        has_editable = True
                    else:
                        td['peut_modifier'] = peut_modifier_fn(td['tache'])
                        if td['peut_modifier']:
                            has_editable = True

                    act_items.append((tx, td['new_poids']))
                    act_statuts.append(st)

                p = sum(w for _, w in act_items)
                ad['taux']   = round(sum(t*w for t,w in act_items)/p, 2) if p else 0.0
                ad['statut'] = _statut_agrege(act_statuts)
                proj_items.append((ad['taux'], ad['new_poids']))

            p = sum(w for _, w in proj_items)
            pjd['taux'] = round(sum(t*w for t,w in proj_items)/p, 2) if p else 0.0
            prog_items.append((pjd['taux'], pjd['new_poids']))

        p = sum(w for _, w in prog_items)
        pd['taux'] = round(sum(t*w for t,w in prog_items)/p, 2) if p else 0.0
        global_items.append((pd['taux'], pd['new_poids']))

    p       = sum(w for _, w in global_items)
    taux_gl = round(sum(t*w for t,w in global_items)/p, 2) if p else 0.0
    return taux_gl, has_editable


def _taux_resp_from_data(data):
    resp = {'programmes': {}, 'projets': {}, 'activites': {}, 'taches': {}}
    for pd in data:
        resp['programmes'][str(pd['programme'].id)] = {'taux': pd['taux']}
        for pjd in pd['projets']:
            resp['projets'][str(pjd['projet'].id)] = {'taux': pjd['taux']}
            for ad in pjd['activites']:
                resp['activites'][str(ad['activite'].id)] = {
                    'taux': ad['taux'], 'statut': ad['statut']}
                for td in ad['taches']:
                    resp['taches'][str(td['tache'].id)] = {
                        'taux': td['taux'], 'statut': td['statut']}
    return resp


def _fmt_periode(debut, fin):
    d = MOIS_COURT.get(debut or '', debut or '')
    f = MOIS_COURT.get(fin   or '', fin   or '')
    if d and f and d != f: return f"{d} - {f}"
    return d or f or ''


# ─────────────────────────────────────────────────────────────────────────────
#  Route principale
# ─────────────────────────────────────────────────────────────────────────────

@suivi_bp.route('/')
@login_required
def index():
    annee = _get_annee()
    if not annee:
        flash("Aucune année active.", 'danger')
        return redirect(url_for('pta.global_pta'))

    trimestre = request.args.get('trimestre', 0, type=int)
    if trimestre not in (0,1,2,3,4):
        trimestre = 0

    # Mode édition : activé si ?edit=1 (fonctionne sur tous les onglets, y compris Global)
    edit_mode = (request.args.get('edit', '0') == '1')

    role = current_user.role

    # ── Suivi map ─────────────────────────────────────────────────────────────
    # Toujours charger le suivi global : statut/taux le plus récent par tâche,
    # indépendamment de l'onglet affiché.
    # Le filtrage par période est géré par _filter_and_renorm, pas par le suivi.
    suivi_map = _load_suivis_global(annee.id)

    # ── Données PTA + permissions ─────────────────────────────────────────────
    data_brut        = []
    titre            = ''
    peut_modifier_fn = None   # par défaut : lecture seule
    peut_editer_pta  = False  # True si le rôle peut potentiellement éditer (indép. de edit_mode)
    directions       = []
    services_list    = []
    services_dir     = []      # services sous la direction courante (rôle direction)
    sel_dir_id       = None
    sel_svc_id       = None
    entite           = None
    niveau           = 'global'  # 'global'|'direction'|'service'

    # ── Rôle SERVICE ──────────────────────────────────────────────────────────
    if role == 'service':
        service = current_user.service
        if not service:
            flash("Aucun service lié à votre compte.", 'danger')
            return redirect(url_for('pta.global_pta'))
        data_brut       = _compute_pta_service(annee, service)
        titre           = f"{service.code} — {service.nom}"
        entite          = service
        niveau          = 'service'
        peut_editer_pta = True
        if edit_mode:
            peut_modifier_fn = True   # toutes les tâches du service sont éditables

    # ── Rôle DIRECTION ────────────────────────────────────────────────────────
    elif role == 'direction':
        direction   = current_user.direction
        if not direction:
            flash("Aucune direction liée à votre compte.", 'danger')
            return redirect(url_for('pta.global_pta'))
        services_dir = Service.query.filter_by(direction_id=direction.id)\
                                    .order_by(Service.nom).all()
        sel_svc_id   = request.args.get('service_id', type=int)
        # Valider que le service appartient bien à la direction
        if sel_svc_id and not any(s.id == sel_svc_id for s in services_dir):
            sel_svc_id = None

        if sel_svc_id:
            # Vue d'un service de la direction — lecture seule pour la direction
            service   = db.session.get(Service, sel_svc_id)
            data_brut = _compute_pta_service(annee, service)
            titre     = f"{service.code} — {service.nom} (lecture)"
            entite    = service
            niveau    = 'service'
            # Pas d'édition pour la direction sur le service d'un autre
        else:
            # Vue totale de la direction (tâches directes + tous ses services)
            data_brut       = _compute_pta_direction(annee, direction)
            titre           = f"{direction.code} — {direction.nom}"
            entite          = direction
            niveau          = 'direction'
            peut_editer_pta  = True
            if edit_mode:
                # La direction édite toutes les tâches de sa direction
                # (filtrées par direction_responsable_id dans _compute_pta_direction,
                #  y compris celles associées à ses services via services_concernes)
                peut_modifier_fn = True

    # ── Rôle ADMIN ────────────────────────────────────────────────────────────
    else:
        directions    = Direction.query.order_by(Direction.nom).all()
        services_list = Service.query.order_by(Service.nom).all()
        sel_svc_id    = request.args.get('service_id', type=int)
        sel_dir_id    = request.args.get('direction_id', type=int)

        if sel_svc_id:
            service   = db.session.get(Service, sel_svc_id)
            if service:
                data_brut = _compute_pta_service(annee, service)
                titre     = f"{service.code} — {service.nom}"
                entite    = service
                niveau    = 'service'
        elif sel_dir_id:
            direction = db.session.get(Direction, sel_dir_id)
            if direction:
                data_brut = _compute_pta_direction(annee, direction)
                titre     = f"{direction.code} — {direction.nom}"
                entite    = direction
                niveau    = 'direction'
        else:
            # Global : tout le PTA
            data_brut = _compute_pta_global(annee)
            titre     = 'Vue globale — tout le PTA'
            niveau    = 'global'
        # admin_editeur : contrôle qualité — peut corriger toutes les tâches
        # admin_lecteur  : lecture seule
        if role == 'admin_editeur':
            peut_editer_pta = True        # affiche le bouton "Faire le suivi"
            if edit_mode:
                peut_modifier_fn = True   # active les champs en mode édition
            # sinon peut_modifier_fn reste None (lecture)

    # ── Filtrage par période + enrichissement ─────────────────────────────────
    data               = _filter_and_renorm(data_brut, trimestre)
    taux_gl, has_edit  = _enrich(data, suivi_map, peut_modifier_fn)
    # peut_modifier_page : vrai dès que edit_mode ET qu'il y a des tâches éditables
    # (valable aussi sur l'onglet Global : l'utilisateur choisit le trimestre via le sélecteur JS)
    peut_modifier_page = has_edit and edit_mode

    return render_template('suivi/index.html',
        annee=annee,
        trimestre=trimestre,
        edit_mode=edit_mode,
        data=data,
        taux_global=taux_gl,
        peut_modifier_page=peut_modifier_page,
        peut_editer_pta=peut_editer_pta,
        titre=titre,
        role=role,
        niveau=niveau,
        entite=entite,
        directions=directions,
        services_list=services_list,
        services_dir=services_dir,
        sel_dir_id=sel_dir_id,
        sel_svc_id=sel_svc_id,
    )


# ─────────────────────────────────────────────────────────────────────────────
#  Route save (AJAX)
# ─────────────────────────────────────────────────────────────────────────────

@suivi_bp.route('/save', methods=['POST'])
@login_required
def save():
    if current_user.role not in ('service', 'direction', 'admin_editeur'):
        return jsonify({'ok': False, 'msg': 'Non autorisé.'}), 403

    annee = _get_annee()
    if not annee:
        return jsonify({'ok': False, 'msg': 'Aucune année active.'}), 400

    payload = request.get_json(silent=True)
    if not payload:
        return jsonify({'ok': False, 'msg': 'Données invalides.'}), 400

    trimestre = payload.get('trimestre')
    # trimestre=0 → vue globale : chaque tâche sera sauvegardée dans son trimestre
    # naturel déterminé par sa période d'exécution (periode_debut)
    if trimestre not in (0, 1, 2, 3, 4):
        return jsonify({'ok': False, 'msg': 'Trimestre invalide.'}), 400

    saved = 0
    for td in payload.get('taches', []):
        try:
            tache_id = int(td['tache_id'])
        except (KeyError, TypeError, ValueError):
            continue

        tache = db.session.get(Tache, tache_id)
        if not tache:
            continue

        if current_user.role == 'service':
            service = current_user.service
            if not service:
                continue
            # Le service peut sauvegarder son suivi si :
            # 1. Il est explicitement associé à la tâche (services_concernes)
            # 2. ET il appartient à la direction responsable de la tâche
            # (une tâche DAF ne peut pas être suivie par un service DTC,
            #  même si ce service apparaît par erreur dans services_concernes)
            ok = any(s.id == service.id and s.direction_id == tache.direction_responsable_id
                     for s in tache.services_concernes)
            if not ok:
                continue
            sid = service.id

        elif current_user.role == 'direction':
            direction = current_user.direction
            if not direction:
                continue
            # La direction peut modifier toutes les tâches de sa direction
            if tache.direction_responsable_id != direction.id:
                continue
            # service_id = le service responsable de la tâche (ou None si tâche directe)
            sid = tache.service_responsable_id

        else:  # admin_editeur — contrôle qualité, accès à toutes les tâches
            # service_id = le responsable naturel de la tâche
            sid = tache.service_responsable_id

        # Vue globale (trimestre=0) : déterminer le trimestre de la tâche
        # Trimestre naturel de la tâche, basé sur sa période d'exécution.
        # On ignore l'onglet courant (trimestre) : le statut/taux est unique par tâche,
        # le filtrage dans les onglets est géré par _filter_and_renorm.
        debut = MOIS_ORDRE.get(tache.periode_debut or '', 0)
        if debut == 0:
            fin = MOIS_ORDRE.get(tache.periode_fin or '', 0)
            debut = fin
        t_tache = ((debut - 1) // 3 + 1) if debut > 0 else 1

        statut = td.get('statut', 'non_execute')
        if statut not in ('execute', 'non_execute', 'en_cours'):
            statut = 'non_execute'

        if statut == 'execute':
            taux = 100.0
        elif statut == 'non_execute':
            taux = 0.0
        else:
            try:
                taux = max(0.0, float(
                    str(td.get('taux','0')).replace(',','.').strip()))
            except (TypeError, ValueError):
                taux = 0.0
            # en_cours : taux doit etre entre 0 exclu et 100 exclu
            nom_t = tache.nom or f'id={tache_id}'
            if taux <= 0.0:
                return jsonify({
                    'ok': False,
                    'msg': (f'Tache \u00ab\u00a0{nom_t}\u00a0\u00bb\u00a0: '
                            f'taux 0\u00a0% incompatible avec le statut \u00ab\u00a0En cours\u00a0\u00bb. '
                            f'Saisissez un taux entre 1\u00a0% et 99\u00a0%, '
                            f'ou choisissez \u00ab\u00a0Non ex\u00e9cut\u00e9e\u00a0\u00bb.')
                }), 400
            if taux >= 100.0:
                return jsonify({
                    'ok': False,
                    'msg': (f'Tache \u00ab\u00a0{nom_t}\u00a0\u00bb\u00a0: '
                            f'taux 100\u00a0% incompatible avec le statut \u00ab\u00a0En cours\u00a0\u00bb. '
                            f'Choisissez \u00ab\u00a0Ex\u00e9cut\u00e9e\u00a0\u00bb si la tache est terminee.')
                }), 400

        observation = str(td.get('observation','') or '').strip()[:3000]

        # Note SQLite : la contrainte UNIQUE ne protège pas les NULL (NULL != NULL en SQL).
        # SQLAlchemy génère correctement 'IS NULL' quand sid is None,
        # ce qui trouve les enregistrements existants avec service_id=NULL.
        if sid is None:
            sv = SuiviTache.query.filter(
                SuiviTache.tache_id   == tache_id,
                SuiviTache.service_id.is_(None),
                SuiviTache.trimestre  == t_tache,
                SuiviTache.annee_id   == annee.id,
            ).first()
        else:
            sv = SuiviTache.query.filter(
                SuiviTache.tache_id   == tache_id,
                SuiviTache.service_id == sid,
                SuiviTache.trimestre  == t_tache,
                SuiviTache.annee_id   == annee.id,
            ).first()

        if sv:
            sv.statut         = statut
            sv.taux_execution = taux
            sv.observation    = observation
            sv.date_maj       = datetime.utcnow()
            sv.modified_by_id = current_user.id
        else:
            sv = SuiviTache(tache_id=tache_id, service_id=sid,
                            trimestre=t_tache, annee_id=annee.id,
                            statut=statut, taux_execution=taux,
                            observation=observation,
                            modified_by_id=current_user.id)
            db.session.add(sv)
        saved += 1

    db.session.commit()
    if saved > 0:
        lbl_tri = f"T{trimestre}" if trimestre else "Global"
        log_audit('suivi_valide', f"Suivi {lbl_tri} — {saved} tâche(s) enregistrée(s)")

    # Recalcul : toujours en mode global (statut le plus récent par tâche)
    suivi_map = _load_suivis_global(annee.id)

    if current_user.role == 'service':
        data_brut = _compute_pta_service(annee, current_user.service)
        pmfn      = True
    elif current_user.role == 'direction':
        direction = current_user.direction
        data_brut = _compute_pta_direction(annee, direction)
        pmfn      = True
    else:  # admin_editeur : recalcul global (la page se rafraîchit 1,5 s après)
        data_brut = _compute_pta_global(annee)
        pmfn      = True

    data              = _filter_and_renorm(data_brut, trimestre)
    taux_gl, _        = _enrich(data, suivi_map, pmfn)
    taux_resp         = _taux_resp_from_data(data)
    taux_resp['global'] = taux_gl

    return jsonify({'ok': True, 'msg': f"{saved} tâche(s) enregistrée(s).", 'taux': taux_resp})


# ─────────────────────────────────────────────────────────────────────────────
#  Export Excel
# ─────────────────────────────────────────────────────────────────────────────

@suivi_bp.route('/export/excel')
@login_required
def export_excel():
    """Exporte le suivi courant en Excel — même style que dirpta/svcpta."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import os as _os
    from flask import current_app

    annee = _get_annee()
    if not annee:
        flash("Aucune année active.", 'danger')
        return redirect(url_for('suivi.index'))

    trimestre  = request.args.get('trimestre', 0, type=int)
    sel_svc_id = request.args.get('service_id', type=int)
    sel_dir_id = request.args.get('direction_id', type=int)
    role       = current_user.role

    suivi_map = _load_suivis_global(annee.id)

    show_service_badge = False   # badge code service dans libellé tâche

    if role == 'service':
        service   = current_user.service
        data_brut = _compute_pta_service(annee, service)
        titre     = f"{service.code} — {service.nom}"
    elif role == 'direction':
        direction = current_user.direction
        if sel_svc_id and Service.query.filter_by(id=sel_svc_id, direction_id=direction.id).first():
            svc       = db.session.get(Service, sel_svc_id)
            data_brut = _compute_pta_service(annee, svc)
            titre     = f"{svc.code} — {svc.nom}"
        else:
            data_brut          = _compute_pta_direction(annee, direction)
            titre              = f"{direction.code} — {direction.nom}"
            show_service_badge = True
    else:
        if sel_svc_id:
            svc       = db.session.get(Service, sel_svc_id)
            data_brut = _compute_pta_service(annee, svc) if svc else []
            titre     = f"{svc.code} — {svc.nom}" if svc else 'Service'
        elif sel_dir_id:
            direction          = db.session.get(Direction, sel_dir_id)
            data_brut          = _compute_pta_direction(annee, direction) if direction else []
            titre              = f"{direction.code} — {direction.nom}" if direction else 'Direction'
            show_service_badge = bool(direction)
        else:
            data_brut          = _compute_pta_global(annee)
            titre              = 'Vue globale'
            show_service_badge = True

    data       = _filter_and_renorm(data_brut, trimestre)
    taux_gl, _ = _enrich(data, suivi_map, None)

    lbl_tri = f"T{trimestre}" if trimestre else "Global"

    # ── Styles ────────────────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = f"Suivi {lbl_tri}"[:31]

    thin  = Side(style='thin')
    med   = Side(style='medium')
    brd   = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr   = Alignment(horizontal='center', vertical='center', wrap_text=True)
    lft   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    rgt   = Alignment(horizontal='right',  vertical='center', wrap_text=True)

    f_prog = PatternFill("solid", fgColor="FFFF99")
    f_proj = PatternFill("solid", fgColor="FFB6C1")
    f_act  = PatternFill("solid", fgColor="D3D3D3")
    f_tch  = PatternFill("solid", fgColor="FFFFFF")
    f_hdr  = PatternFill("solid", fgColor="FF69B4")
    f_tot  = PatternFill("solid", fgColor="1F6B35")   # vert foncé
    f_obj  = PatternFill("solid", fgColor="FFFACD")
    f_tit  = PatternFill("solid", fgColor="D1F0DA")   # vert clair
    NCOLS  = 7   # A–G (sans Responsable)

    STATUTS = {'execute': 'Exécutée', 'en_cours': 'En cours', 'non_execute': 'Non exécutée'}

    def wr(vals, fill, bold, rn, aligns=None):
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=rn, column=col, value=v)
            c.fill = fill
            c.font = Font(bold=bold, size=9)
            c.border = brd
            al = aligns[col-1] if aligns else (lft if col == 2 else ctr)
            c.alignment = al
        return rn

    # ── Lignes 1-3 : En-tête institutionnel ───────────────────────────────────
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 8
    ws.row_dimensions[3].height = 22
    ws.merge_cells('A1:B3')
    ws.merge_cells('C1:E3')
    ws.merge_cells('F1:G3')
    c = ws['C1']
    c.value = "BP 02 Adja-Ouèrè\nTél : +229 01 61 91 96 12\nEmail : contact.adjaouere@mairie.bj"
    c.font = Font(bold=True, size=9)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Images bandeau + logo
    try:
        from openpyxl.drawing.image import Image as _XLImg
        from PIL import Image as _PILImg
        import io as _imgio
        _simg = _os.path.join(current_app.root_path, 'static', 'img')
        _H = 52
        _band = _os.path.join(_simg, 'bandeau.png')
        _logo = _os.path.join(_simg, 'logo_commune.png')
        if _os.path.exists(_band):
            _buf = _imgio.BytesIO()
            with _PILImg.open(_band) as _p:
                _ow, _oh = _p.size; _p.save(_buf, 'PNG')
            _buf.seek(0)
            _i = _XLImg(_buf); _i.height = _H; _i.width = int(_ow * _H / _oh)
            ws.add_image(_i, 'A1')
        if _os.path.exists(_logo):
            _buf2 = _imgio.BytesIO()
            with _PILImg.open(_logo) as _p2:
                _ow2, _oh2 = _p2.size; _p2.save(_buf2, 'PNG')
            _buf2.seek(0)
            _i2 = _XLImg(_buf2); _lw = int(_ow2 * _H / _oh2)
            _i2.height = _H; _i2.width = _lw
            try:
                from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
                from openpyxl.drawing.xdr import XDRPositiveSize2D
                _a = OneCellAnchor()
                _a._from = AnchorMarker(col=6, colOff=-int(_lw * 9525), row=0, rowOff=0)
                _a.ext = XDRPositiveSize2D(int(_lw * 9525), int(_H * 9525))
                _i2.anchor = _a; ws.add_image(_i2)
            except Exception:
                ws.add_image(_i2, 'F1')
    except Exception:
        pass

    # ── Ligne 4 : Titre ───────────────────────────────────────────────────────
    ws.merge_cells('A4:G4')
    _tri_lbl = f"Trimestre {trimestre}" if trimestre else "Tous trimestres"
    ws['A4'].value = (f"SUIVI D'EXÉCUTION DU PTA — Exercice {annee.annee}"
                      f"   |   {titre}   |   {_tri_lbl}"
                      f"   |   Taux : {taux_gl:.2f}%")
    ws['A4'].font = Font(bold=True, size=12)
    ws['A4'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A4'].fill = f_tit
    for _tc in range(1, NCOLS+1):
        ws.cell(row=4, column=_tc).border = Border(
            left=med if _tc==1 else Side(style=None),
            right=med if _tc==NCOLS else Side(style=None),
            top=med, bottom=med)
    ws.row_dimensions[4].height = 28

    # ── Ligne 5 : Objectif général ────────────────────────────────────────────
    if annee.objectif_general:
        ws.merge_cells('A5:G5')
        c = ws['A5']
        c.value = f"Objectif général : {annee.objectif_general}"
        c.font = Font(bold=True, italic=True, size=9); c.alignment = lft
        start_row = 6
    else:
        start_row = 5

    # ── En-têtes tableau ──────────────────────────────────────────────────────
    hdrs = ['Code', 'Libellé / Activité / Tâche', 'Période',
            'Poids (%)', 'Statut', 'Taux (%)', 'Observations / Difficultés']
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=start_row, column=ci, value=h)
        c.fill = f_hdr; c.font = Font(bold=True, size=9, color="000000")
        c.alignment = ctr; c.border = brd
    ws.row_dimensions[start_row].height = 20

    row = start_row + 1

    # ── Données ───────────────────────────────────────────────────────────────
    for pd in data:
        prog = pd['programme']
        if prog.objectif_specifique:
            ws.merge_cells(f'A{row}:G{row}')
            c = ws.cell(row=row, column=1,
                value=f"Objectif {pd['code']}/Résultat {pd['code']} : {prog.objectif_specifique}")
            c.font = Font(bold=True, italic=True, size=9)
            c.fill = f_obj; c.alignment = lft; c.border = brd
            row += 1

        wr([pd['code'], f"Programme {pd['code']} : {prog.nom}", '',
            f"{pd['new_poids']:.2f}%", '', f"{pd['taux']:.2f}%", ''],
           f_prog, True, row)
        row += 1

        for pjd in pd['projets']:
            proj = pjd['projet']
            wr([pjd['code'], f"Projet {pjd['code']} : {proj.nom}", '',
                f"{pjd['new_poids']:.2f}%", '', f"{pjd['taux']:.2f}%", ''],
               f_proj, True, row)
            row += 1

            for ad in pjd['activites']:
                act    = ad['activite']
                per_a  = _fmt_periode(act.periode_debut, act.periode_fin)
                st_lbl = STATUTS.get(ad['statut'] or 'non_execute', '—')
                wr([ad['code'], act.nom, per_a,
                    f"{ad['new_poids']:.2f}%", st_lbl, f"{ad['taux']:.2f}%", ''],
                   f_act, True, row)
                row += 1

                for td in ad['taches']:
                    t      = td['tache']
                    per_t  = _fmt_periode(t.periode_debut, t.periode_fin)
                    st_lbl = STATUTS.get(td['statut'] or 'non_execute', '—')
                    obs    = td['suivi'].observation if td['suivi'] and td['suivi'].observation else ''
                    # Code service en préfixe du libellé si vue direction/global
                    nom_t  = f"  {t.nom}"
                    if show_service_badge and t.service_responsable:
                        nom_t = f"  [{t.service_responsable.code}] {t.nom}"
                    wr([td['num'], nom_t, per_t,
                        f"{td['new_poids']:.2f}%", st_lbl, f"{td['taux']:.2f}%", obs],
                       f_tch, False, row,
                       aligns=[ctr, lft, ctr, ctr, ctr, ctr, lft])
                    row += 1

    # ── Total général ─────────────────────────────────────────────────────────
    for ci, val in enumerate(['', 'TOTAL GÉNÉRAL', '', '100%', '',
                               f"{taux_gl:.2f}%", ''], 1):
        c = ws.cell(row=row, column=ci, value=val)
        c.fill = f_tot
        c.font = Font(bold=True, size=9, color="FFFFFF")
        c.border = brd; c.alignment = ctr
    row += 1

    # ── Pied ──────────────────────────────────────────────────────────────────
    _date_str = datetime.now().strftime('%d/%m/%Y à %H:%M')
    ws.merge_cells(f'A{row}:C{row}')
    c = ws.cell(row=row, column=1, value=f"Exporté le {_date_str}")
    c.font = Font(bold=True, size=8); c.alignment = lft
    ws.merge_cells(f'D{row}:G{row}')
    c = ws.cell(row=row, column=4,
                value="Direction du Développement Local et de la Planification (DDLP)")
    c.font = Font(bold=True, size=8); c.alignment = rgt

    # ── Largeurs colonnes (A–G) ───────────────────────────────────────────────
    for ci, w in enumerate([8, 46, 12, 9, 16, 9, 42], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = f'A{start_row + 1}'

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f"Suivi_{lbl_tri}_{annee.annee}_{titre[:20].replace(' ','_').replace('/','_')}.xlsx"
    return send_file(buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name=fname)


# ─────────────────────────────────────────────────────────────────────────────
#  Vue impression / PDF
# ─────────────────────────────────────────────────────────────────────────────

@suivi_bp.route('/print')
@login_required
def print_view():
    """Page optimisée pour impression / export PDF (Ctrl+P)."""
    annee = _get_annee()
    if not annee:
        flash("Aucune année active.", 'danger')
        return redirect(url_for('suivi.index'))

    trimestre  = request.args.get('trimestre', 0, type=int)
    sel_svc_id = request.args.get('service_id', type=int)
    sel_dir_id = request.args.get('direction_id', type=int)
    role       = current_user.role

    suivi_map = _load_suivis_global(annee.id)

    show_service_badge = False   # badge code service dans libellé tâche

    if role == 'service':
        service   = current_user.service
        data_brut = _compute_pta_service(annee, service)
        titre     = f"{service.code} — {service.nom}"
    elif role == 'direction':
        direction = current_user.direction
        if sel_svc_id and Service.query.filter_by(id=sel_svc_id, direction_id=direction.id).first():
            svc       = db.session.get(Service, sel_svc_id)
            data_brut = _compute_pta_service(annee, svc)
            titre     = f"{svc.code} — {svc.nom}"
        else:
            data_brut          = _compute_pta_direction(annee, direction)
            titre              = f"{direction.code} — {direction.nom}"
            show_service_badge = True   # vue direction globale : montrer code service
    else:
        if sel_svc_id:
            svc       = db.session.get(Service, sel_svc_id)
            data_brut = _compute_pta_service(annee, svc) if svc else []
            titre     = f"{svc.code} — {svc.nom}" if svc else ''
        elif sel_dir_id:
            direction          = db.session.get(Direction, sel_dir_id)
            data_brut          = _compute_pta_direction(annee, direction) if direction else []
            titre              = f"{direction.code} — {direction.nom}" if direction else ''
            show_service_badge = bool(direction)
        else:
            data_brut          = _compute_pta_global(annee)
            titre              = 'Vue globale — tout le PTA'
            show_service_badge = True   # vue globale : montrer code service dans les tâches

    data    = _filter_and_renorm(data_brut, trimestre)
    taux_gl, _ = _enrich(data, suivi_map, None)

    return render_template('suivi/print.html',
        annee=annee, trimestre=trimestre, data=data,
        taux_global=taux_gl, titre=titre,
        show_service_badge=show_service_badge,
        fmt_periode=_fmt_periode,
        now_str=datetime.now().strftime('%d/%m/%Y %H:%M'),
    )
