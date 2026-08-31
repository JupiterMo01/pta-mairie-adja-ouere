"""
Script d'import du PTA depuis un fichier Excel.
Usage : python import_pta.py [chemin_vers_fichier.xlsx] [annee]

  python import_pta.py "C:/Users/HP/Desktop/Classeur1.xlsx" 2026

Si aucun argument n'est fourni, utilise les valeurs par défaut ci-dessous.
"""

import sys
import os

# ─── Paramètres par défaut ────────────────────────────────────────────────────
EXCEL_PATH = r"C:\Users\HP\Desktop\Classeur1.xlsx"
ANNEE_CIBLE = 2026
SHEET_NAME  = 'PTA 2026'
PREMIERE_LIGNE_DONNEES = 14   # on commence tôt, les lignes vides sont ignorées

# ─── Mapping abréviations → noms complets des mois ───────────────────────────
MOIS_MAP = {
    'janv':    'Janvier',
    'jan':     'Janvier',
    'fév':     'Février',
    'fev':     'Février',
    'fevr':    'Février',
    'mars':    'Mars',
    'avr':     'Avril',
    'avril':   'Avril',
    'mai':     'Mai',
    'juin':    'Juin',
    'juil':    'Juillet',
    'juillet': 'Juillet',
    'août':    'Août',
    'aout':    'Août',
    'sep':     'Septembre',
    'sept':    'Septembre',
    'oct':     'Octobre',
    'nov':     'Novembre',
    'déc':     'Décembre',
    'dec':     'Décembre',
}


def _normaliser_mois(s):
    """Convertit une abréviation de mois en nom complet, ou retourne None."""
    if not s:
        return None
    key = s.strip().lower().rstrip('.')
    return MOIS_MAP.get(key)


def _parse_periode(val):
    """
    Parse une cellule période comme :
      "Janv - Déc", "Janv-Juin", "Avril-Juin", "Oct - Déc",
      "Janv - Mars", "Juillet - Sept", "Mars-Juin",
      "Janv", "Fév", "Mai", …
    Retourne (debut, fin) ou (debut, debut) pour un seul mois.
    """
    if not val:
        return None, None
    s = str(val).strip()
    # séparer sur " - " ou "-" (avec ou sans espaces)
    for sep in (' - ', ' -', '- ', '-'):
        if sep in s:
            parts = s.split(sep, 1)
            debut = _normaliser_mois(parts[0].strip())
            fin   = _normaliser_mois(parts[1].strip())
            return debut, fin
    # un seul mois
    m = _normaliser_mois(s)
    return m, m


def _parse_montant(val):
    """Parse un montant Excel (float, int, str avec espaces/virgules)."""
    if val is None or val == '':
        return 0.0
    s = str(val).strip().replace('\xa0', '').replace(' ', '')
    if not s or s in ('-', '—', 'néant', 'neant', 'n/a'):
        return 0.0
    # format français : virgule comme décimale
    if ',' in s:
        s = s.replace('.', '').replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return 0.0


def _niveau_code(code_str):
    """
    Retourne le niveau hiérarchique à partir du code :
      "1"       → 1  (Programme)
      "1.1"     → 2  (Projet)
      "1.1.1"   → 3  (Activité)
      "1.1.1.1" → 4  (Tâche)
    Retourne None si le code est invalide.
    """
    if not code_str:
        return None
    s = str(code_str).strip().rstrip('.')
    if not s:
        return None
    parts = [p for p in s.split('.') if p.strip()]
    # Chaque partie doit être un entier
    for p in parts:
        try:
            int(p.strip())
        except ValueError:
            return None
    return len(parts)


def _code_parent(code_str):
    """Retourne le code du parent : "1.2.3" → "1.2" """
    parts = str(code_str).strip().rstrip('.').split('.')
    return '.'.join(parts[:-1])


def _dernier_numero(code_str):
    """Retourne le dernier segment numérique : "1.2.3" → 3"""
    parts = [p.strip() for p in str(code_str).strip().rstrip('.').split('.') if p.strip()]
    return int(parts[-1])


# ─── Couleur jaune = Programme ────────────────────────────────────────────────
YELLOW_FILL = 'FFFFFF00'

def _est_programme(cell_b, cell_d):
    """
    Un Programme est une ligne jaune OU dont le code B est un entier pur (pas de point).
    """
    # Vérification couleur de fond
    try:
        fill = cell_b.fill
        if fill and fill.fgColor:
            rgb = fill.fgColor.rgb
            if rgb and rgb.upper() in (YELLOW_FILL, 'FF' + YELLOW_FILL):
                return True
    except Exception:
        pass
    # Vérification par structure du code
    val = cell_b.value
    if val is None:
        return False
    try:
        int(str(val).strip())
        return True    # code entier pur = Programme
    except ValueError:
        return False


# ─── Lookup services/directions ───────────────────────────────────────────────

_cache_svc = {}   # code → Service
_cache_dir = {}   # code → Direction

def _init_caches(app_models):
    from models import Service, Direction
    for s in Service.query.all():
        _cache_svc[s.code.upper()] = s
    for d in Direction.query.all():
        _cache_dir[d.code.upper()] = d


def _trouver_org(token):
    """
    Cherche d'abord dans les services, puis dans les directions.
    Retourne (service_ou_none, direction_ou_none).
    """
    if not token:
        return None, None
    key = token.strip().upper()
    svc = _cache_svc.get(key)
    if svc:
        return svc, None
    dir_ = _cache_dir.get(key)
    if dir_:
        return None, dir_
    return None, None


def _tokens(val):
    """
    Découpe une valeur de cellule en tokens individuels.
    Séparateurs : virgule ET slash (ex: "SBFC/DAAF,SE" → ["SBFC","DAAF","SE"]).
    Nettoie les espaces et le slash de tête.
    """
    if not val:
        return []
    result = []
    for part in str(val).strip().lstrip('/').split(','):
        for sub in part.split('/'):
            t = sub.strip()
            if t:
                result.append(t)
    return result


def _parse_responsable(val):
    """
    Parse la colonne M (responsable). Exemples : 'DDLP', 'SBFC/DAAF', 'SVEMI/DST', '/DAAF'.
    Retourne (service_ou_none, direction_ou_none, texte_non_reconnu).
    """
    svc_resp = dir_resp = None
    unrec = []
    for t in _tokens(val):
        svc, dir_ = _trouver_org(t)
        if svc and svc_resp is None:
            svc_resp = svc
        elif dir_ and dir_resp is None:
            dir_resp = dir_
        elif not svc and not dir_:
            unrec.append(t)
    return svc_resp, dir_resp, unrec


def _parse_structures(val):
    """
    Parse la colonne N (structures associées).
    Retourne (services[], directions[], textes_non_reconnus[]).
    """
    if not val:
        return [], [], []
    svcs, dirs, unrec = [], [], []
    for t in _tokens(val):
        svc, dir_ = _trouver_org(t)
        if svc and svc not in svcs:
            svcs.append(svc)
        elif dir_ and dir_ not in dirs:
            dirs.append(dir_)
        elif not svc and not dir_ and t not in unrec:
            unrec.append(t)
    return svcs, dirs, unrec


# ─── Import principal ─────────────────────────────────────────────────────────

def run_import(excel_path, annee_val):
    import openpyxl
    from app import create_app
    from models import (db, Annee, Programme, Projet, Activite, Tache,
                        StructureExterne)

    flask_app = create_app()

    with flask_app.app_context():
        _init_caches(None)

        # Trouver l'année cible
        annee = Annee.query.filter_by(annee=annee_val).first()
        if not annee:
            print(f"ERREUR : L'année {annee_val} n'existe pas dans la base.")
            print("Années disponibles :", [a.annee for a in Annee.query.all()])
            sys.exit(1)

        # Vérifier que le PTA est vide
        nb_progs = Programme.query.filter_by(annee_id=annee.id).count()
        if nb_progs > 0:
            print(f"ERREUR : Le PTA {annee_val} contient déjà {nb_progs} programme(s).")
            print("Réinitialisez-le d'abord depuis l'interface (Administration › PTA Global › Réinitialiser).")
            sys.exit(1)

        # Charger le classeur
        print(f"Ouverture de : {excel_path}")
        wb = openpyxl.load_workbook(excel_path, data_only=True)

        # Choisir la feuille
        sheet = None
        for name in wb.sheetnames:
            if str(annee_val) in name or name.strip().lower() == SHEET_NAME.lower():
                sheet = wb[name]
                break
        if sheet is None:
            sheet = wb.active
        print(f"Feuille utilisée : {sheet.title}  ({sheet.max_row} lignes)")

        # Compteurs
        nb_prog = nb_proj = nb_act = nb_tach = 0
        nb_skip = 0

        # Contexte courant
        cur_prog  = None   # objet Programme en cours
        cur_proj  = None   # objet Projet en cours
        cur_act   = None   # objet Activite en cours

        # Numérotation locale par parent
        # On fait confiance aux numéros du fichier Excel.

        for row_idx in range(PREMIERE_LIGNE_DONNEES, sheet.max_row + 1):
            cell_b = sheet.cell(row=row_idx, column=2)   # Code
            cell_c = sheet.cell(row=row_idx, column=3)   # Désignation
            cell_d = sheet.cell(row=row_idx, column=4)   # Imputation / Objectif
            cell_e = sheet.cell(row=row_idx, column=5)   # RP
            cell_f = sheet.cell(row=row_idx, column=6)   # FADeC Aff
            cell_g = sheet.cell(row=row_idx, column=7)   # FADeC Non Aff
            cell_h = sheet.cell(row=row_idx, column=8)   # Autres Part
            cell_i = sheet.cell(row=row_idx, column=9)   # Autres Fonds
            cell_k = sheet.cell(row=row_idx, column=11)  # Période
            cell_l = sheet.cell(row=row_idx, column=12)  # Poids
            cell_m = sheet.cell(row=row_idx, column=13)  # Responsable
            cell_n = sheet.cell(row=row_idx, column=14)  # Structures associées
            cell_o = sheet.cell(row=row_idx, column=15)  # Mode exécution

            code_val = cell_b.value
            desig    = str(cell_c.value).strip() if cell_c.value else ''

            # Ligne vide → ignorer
            if code_val is None and not desig:
                nb_skip += 1
                continue

            # ── PROGRAMME ──────────────────────────────────────────────────────
            if _est_programme(cell_b, cell_d):
                try:
                    numero = int(str(code_val).strip())
                except (ValueError, TypeError):
                    nb_skip += 1
                    continue

                # Nom du programme = colonne D (objectif/description)
                nom_prog = str(cell_d.value).strip() if cell_d.value else desig
                if not nom_prog:
                    nom_prog = desig or f'Programme {numero}'

                poids = _parse_montant(cell_l.value)

                cur_prog = Programme(
                    annee_id=annee.id,
                    numero=numero,
                    nom=nom_prog,
                    poids=poids,
                    observations=str(cell_o.value).strip() if cell_o.value else None,
                )
                db.session.add(cur_prog)
                db.session.flush()   # obtenir l'id
                cur_proj = None
                cur_act  = None
                nb_prog += 1
                print(f"  PROG {numero}: {nom_prog[:60]}")
                continue

            # Pour tout le reste, on a besoin du code
            if code_val is None:
                nb_skip += 1
                continue

            code_str = str(code_val).strip()
            niveau = _niveau_code(code_str)

            if niveau is None:
                nb_skip += 1
                continue

            # ── PROJET ─────────────────────────────────────────────────────────
            if niveau == 2:
                if cur_prog is None:
                    print(f"  WARN ligne {row_idx}: Projet {code_str} sans Programme courant, ignoré")
                    nb_skip += 1
                    continue
                numero = _dernier_numero(code_str)
                nom_proj = desig or f'Projet {code_str}'
                poids = _parse_montant(cell_l.value)

                cur_proj = Projet(
                    programme_id=cur_prog.id,
                    numero=numero,
                    nom=nom_proj,
                    poids=poids,
                    observations=str(cell_o.value).strip() if cell_o.value else None,
                )
                db.session.add(cur_proj)
                db.session.flush()
                cur_act = None
                nb_proj += 1
                continue

            # ── ACTIVITÉ ───────────────────────────────────────────────────────
            if niveau == 3:
                if cur_proj is None:
                    print(f"  WARN ligne {row_idx}: Activité {code_str} sans Projet courant, ignorée")
                    nb_skip += 1
                    continue
                numero = _dernier_numero(code_str)
                nom_act = desig or f'Activité {code_str}'
                poids = _parse_montant(cell_l.value)

                imputation = str(cell_d.value).strip() if cell_d.value else None
                if imputation and imputation.lower() in ('néant', 'neant', '-', ''):
                    imputation = None

                rp  = _parse_montant(cell_e.value)
                fa  = _parse_montant(cell_f.value)
                fn  = _parse_montant(cell_g.value)
                ap  = _parse_montant(cell_h.value)
                af  = _parse_montant(cell_i.value)

                debut, fin = _parse_periode(cell_k.value)
                mode = str(cell_o.value).strip() if cell_o.value else 'Direct'
                if not mode:
                    mode = 'Direct'

                # Responsable (colonne M)
                svc_resp_a, dir_resp_a, unrec_m = _parse_responsable(cell_m.value)
                # Activité : le responsable est une direction ; si seul un service
                # a été trouvé, utiliser sa direction parente
                if svc_resp_a and not dir_resp_a:
                    dir_resp_a = svc_resp_a.direction

                # Structures associées (colonne N)
                svcs_ass, dirs_ass, unrec_n = _parse_structures(cell_n.value)

                # Tout ce qui n'est pas reconnu → acteurs_externes (texte libre)
                tous_unrec = [t for t in unrec_m + unrec_n if t]
                acteurs_txt = ', '.join(tous_unrec) if tous_unrec else None

                cur_act = Activite(
                    projet_id=cur_proj.id,
                    numero=numero,
                    nom=nom_act,
                    poids=poids,
                    imputation_budgetaire=imputation,
                    ressources_propres=rp,
                    fadec_affecte=fa,
                    fadec_non_affecte=fn,
                    autres_partenaires=ap,
                    autres_fonds=af,
                    periode_debut=debut,
                    periode_fin=fin,
                    mode_execution=mode,
                    direction_responsable_id=dir_resp_a.id if dir_resp_a else None,
                    acteurs_externes=acteurs_txt,
                )
                db.session.add(cur_act)
                db.session.flush()

                if svcs_ass:
                    cur_act.services_intervenants = svcs_ass
                if dirs_ass:
                    cur_act.directions_associees = dirs_ass

                nb_act += 1
                continue

            # ── TÂCHE ──────────────────────────────────────────────────────────
            if niveau == 4:
                if cur_act is None:
                    print(f"  WARN ligne {row_idx}: Tâche {code_str} sans Activité courante, ignorée")
                    nb_skip += 1
                    continue
                numero = _dernier_numero(code_str)
                nom_tach = desig or f'Tâche {code_str}'
                poids = _parse_montant(cell_l.value)

                imputation = str(cell_d.value).strip() if cell_d.value else None
                if imputation and imputation.lower() in ('néant', 'neant', '-', ''):
                    imputation = None

                rp  = _parse_montant(cell_e.value)
                fa  = _parse_montant(cell_f.value)
                fn  = _parse_montant(cell_g.value)
                ap  = _parse_montant(cell_h.value)
                af  = _parse_montant(cell_i.value)

                debut, fin = _parse_periode(cell_k.value)
                mode = str(cell_o.value).strip() if cell_o.value else 'Direct'
                if not mode:
                    mode = 'Direct'

                # Responsable tâche (colonne M)
                svc_resp, dir_resp_t, unrec_m = _parse_responsable(cell_m.value)

                # Structures associées (colonne N)
                svcs_ass, dirs_ass, unrec_n = _parse_structures(cell_n.value)

                tous_unrec = [t for t in unrec_m + unrec_n if t]
                acteurs_txt = ', '.join(tous_unrec) if tous_unrec else None

                tache = Tache(
                    activite_id=cur_act.id,
                    numero=numero,
                    ordre=numero,
                    nom=nom_tach,
                    poids=poids,
                    imputation_budgetaire=imputation,
                    ressources_propres=rp,
                    fadec_affecte=fa,
                    fadec_non_affecte=fn,
                    autres_partenaires=ap,
                    autres_fonds=af,
                    mode_execution=mode,
                    periode_debut=debut,
                    periode_fin=fin,
                    service_responsable_id=svc_resp.id if svc_resp else None,
                    direction_responsable_id=dir_resp_t.id if dir_resp_t else None,
                    acteurs_externes=acteurs_txt,
                )
                db.session.add(tache)
                db.session.flush()

                if svcs_ass:
                    tache.services_concernes = svcs_ass
                if dirs_ass:
                    tache.directions_associees = dirs_ass

                nb_tach += 1
                continue

            # Niveau inconnu
            nb_skip += 1

        # Commit final
        db.session.commit()
        print()
        print("=" * 55)
        print(f"Import terminé pour le PTA {annee_val}")
        print(f"  Programmes : {nb_prog}")
        print(f"  Projets    : {nb_proj}")
        print(f"  Activités  : {nb_act}")
        print(f"  Tâches     : {nb_tach}")
        print(f"  Ignorées   : {nb_skip}")
        print("=" * 55)


# ─── Point d'entrée ───────────────────────────────────────────────────────────

if __name__ == '__main__':
    path  = sys.argv[1] if len(sys.argv) > 1 else EXCEL_PATH
    annee = int(sys.argv[2]) if len(sys.argv) > 2 else ANNEE_CIBLE

    if not os.path.exists(path):
        print(f"ERREUR : Fichier introuvable : {path}")
        sys.exit(1)

    run_import(path, annee)
