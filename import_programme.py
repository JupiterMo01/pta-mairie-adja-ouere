"""
Import d'UN SEUL programme depuis le fichier Excel du PTA.
Usage : python import_programme.py <numero_programme> [chemin_excel] [annee]

Exemples :
  python import_programme.py 3
  python import_programme.py 3 "C:/Users/HP/Desktop/Classeur1.xlsx" 2026
"""

import sys
import os

# ─── Paramètres par défaut ────────────────────────────────────────────────────
EXCEL_PATH  = r"C:\Users\HP\Desktop\Classeur1.xlsx"
ANNEE_CIBLE = 2026
SHEET_NAME  = 'PTA 2026'
PREMIERE_LIGNE_DONNEES = 14

# ─── Mapping mois ────────────────────────────────────────────────────────────
MOIS_MAP = {
    'janv': 'Janvier', 'jan': 'Janvier',
    'fév': 'Février', 'fev': 'Février', 'fevr': 'Février',
    'mars': 'Mars',
    'avr': 'Avril', 'avril': 'Avril',
    'mai': 'Mai',
    'juin': 'Juin',
    'juil': 'Juillet', 'juillet': 'Juillet',
    'août': 'Août', 'aout': 'Août',
    'sep': 'Septembre', 'sept': 'Septembre',
    'oct': 'Octobre',
    'nov': 'Novembre',
    'déc': 'Décembre', 'dec': 'Décembre',
}


def _normaliser_mois(s):
    if not s:
        return None
    key = s.strip().lower().rstrip('.')
    return MOIS_MAP.get(key)


def _parse_periode(val):
    if not val:
        return None, None
    s = str(val).strip()
    for sep in (' - ', ' -', '- ', '-'):
        if sep in s:
            parts = s.split(sep, 1)
            return _normaliser_mois(parts[0].strip()), _normaliser_mois(parts[1].strip())
    m = _normaliser_mois(s)
    return m, m


def _parse_montant(val):
    if val is None or val == '':
        return 0.0
    s = str(val).strip().replace('\xa0', '').replace(' ', '')
    if not s or s in ('-', '—', 'néant', 'neant', 'n/a'):
        return 0.0
    if ',' in s:
        s = s.replace('.', '').replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return 0.0


def _niveau_code(code_str):
    if not code_str:
        return None
    s = str(code_str).strip().rstrip('.')
    if not s:
        return None
    parts = [p for p in s.split('.') if p.strip()]
    for p in parts:
        try:
            int(p.strip())
        except ValueError:
            return None
    return len(parts)


def _dernier_numero(code_str):
    parts = [p.strip() for p in str(code_str).strip().rstrip('.').split('.') if p.strip()]
    return int(parts[-1])


YELLOW_FILL = 'FFFFFF00'


def _est_programme(cell_b, cell_d):
    try:
        fill = cell_b.fill
        if fill and fill.fgColor:
            rgb = fill.fgColor.rgb
            if rgb and rgb.upper() in (YELLOW_FILL, 'FF' + YELLOW_FILL):
                return True
    except Exception:
        pass
    val = cell_b.value
    if val is None:
        return False
    try:
        int(str(val).strip())
        return True
    except ValueError:
        return False


_cache_svc = {}
_cache_dir = {}


def _init_caches():
    from models import Service, Direction
    for s in Service.query.all():
        _cache_svc[s.code.upper()] = s
    for d in Direction.query.all():
        _cache_dir[d.code.upper()] = d


def _trouver_org(token):
    if not token:
        return None, None
    key = token.strip().upper()
    svc = _cache_svc.get(key)
    if svc:
        return svc, None
    dir_ = _cache_dir.get(key)
    if dir_:
        return None, dir_
    return None, None


def _tokens(val):
    if not val:
        return []
    result = []
    for part in str(val).strip().lstrip('/').split(','):
        for sub in part.split('/'):
            t = sub.strip()
            if t:
                result.append(t)
    return result


def _parse_responsable(val):
    svc_resp = dir_resp = None
    unrec = []
    for t in _tokens(val):
        svc, dir_ = _trouver_org(t)
        if svc and svc_resp is None:
            svc_resp = svc
        elif dir_ and dir_resp is None:
            dir_resp = dir_
        elif not svc and not dir_:
            unrec.append(t)
    return svc_resp, dir_resp, unrec


def _parse_structures(val):
    if not val:
        return [], [], []
    svcs, dirs, unrec = [], [], []
    for t in _tokens(val):
        svc, dir_ = _trouver_org(t)
        if svc and svc not in svcs:
            svcs.append(svc)
        elif dir_ and dir_ not in dirs:
            dirs.append(dir_)
        elif not svc and not dir_ and t not in unrec:
            unrec.append(t)
    return svcs, dirs, unrec


# ─── Import principal ─────────────────────────────────────────────────────────

def run_import(prog_cible, excel_path, annee_val):
    import openpyxl
    from app import create_app
    from models import db, Annee, Programme, Projet, Activite, Tache

    flask_app = create_app()

    with flask_app.app_context():
        _init_caches()

        # Trouver l'année cible
        annee = Annee.query.filter_by(annee=annee_val).first()
        if not annee:
            print(f"ERREUR : L'année {annee_val} n'existe pas.")
            sys.exit(1)

        # Vérifier que ce programme n'existe pas déjà
        existant = Programme.query.filter_by(annee_id=annee.id, numero=prog_cible).first()
        if existant:
            print(f"ERREUR : Le programme {prog_cible} existe déjà dans le PTA {annee_val}.")
            print("Supprimez-le d'abord depuis l'interface avant de ré-importer.")
            sys.exit(1)

        print(f"Ouverture de : {excel_path}")
        wb = openpyxl.load_workbook(excel_path, data_only=True)

        sheet = None
        for name in wb.sheetnames:
            if str(annee_val) in name or name.strip().lower() == SHEET_NAME.lower():
                sheet = wb[name]
                break
        if sheet is None:
            sheet = wb.active
        print(f"Feuille utilisée : {sheet.title}  ({sheet.max_row} lignes)")
        print(f"Import du programme {prog_cible} uniquement...\n")

        nb_prog = nb_proj = nb_act = nb_tach = nb_skip = 0

        cur_prog = None
        cur_proj = None
        cur_act  = None
        dans_cible = False   # True quand on est dans le programme ciblé

        for row_idx in range(PREMIERE_LIGNE_DONNEES, sheet.max_row + 1):
            cell_b = sheet.cell(row=row_idx, column=2)
            cell_c = sheet.cell(row=row_idx, column=3)
            cell_d = sheet.cell(row=row_idx, column=4)
            cell_e = sheet.cell(row=row_idx, column=5)
            cell_f = sheet.cell(row=row_idx, column=6)
            cell_g = sheet.cell(row=row_idx, column=7)
            cell_h = sheet.cell(row=row_idx, column=8)
            cell_i = sheet.cell(row=row_idx, column=9)
            cell_k = sheet.cell(row=row_idx, column=11)
            cell_l = sheet.cell(row=row_idx, column=12)
            cell_m = sheet.cell(row=row_idx, column=13)
            cell_n = sheet.cell(row=row_idx, column=14)
            cell_o = sheet.cell(row=row_idx, column=15)

            code_val = cell_b.value
            desig    = str(cell_c.value).strip() if cell_c.value else ''

            if code_val is None and not desig:
                nb_skip += 1
                continue

            # ── PROGRAMME ──────────────────────────────────────────────────────
            if _est_programme(cell_b, cell_d):
                try:
                    numero = int(str(code_val).strip())
                except (ValueError, TypeError):
                    nb_skip += 1
                    continue

                if numero == prog_cible:
                    nom_prog = str(cell_d.value).strip() if cell_d.value else desig
                    if not nom_prog:
                        nom_prog = desig or f'Programme {numero}'
                    poids = _parse_montant(cell_l.value)
                    cur_prog = Programme(
                        annee_id=annee.id,
                        numero=numero,
                        nom=nom_prog,
                        poids=poids,
                    )
                    db.session.add(cur_prog)
                    db.session.flush()
                    cur_proj = None
                    cur_act  = None
                    dans_cible = True
                    nb_prog += 1
                    print(f"  PROG {numero}: {nom_prog[:70]}")
                else:
                    # Autre programme : on sort de la cible si on y était
                    if dans_cible and numero > prog_cible:
                        print(f"\n  → Programme {numero} rencontré, fin de l'import du programme {prog_cible}.")
                        break
                    cur_prog = None
                    cur_proj = None
                    cur_act  = None
                    dans_cible = False
                    nb_skip += 1
                continue

            # Si on n'est pas dans le programme cible, ignorer silencieusement
            if not dans_cible:
                nb_skip += 1
                continue

            if code_val is None:
                nb_skip += 1
                continue

            code_str = str(code_val).strip()
            niveau = _niveau_code(code_str)
            if niveau is None:
                nb_skip += 1
                continue

            # ── PROJET ─────────────────────────────────────────────────────────
            if niveau == 2:
                if cur_prog is None:
                    nb_skip += 1
                    continue
                numero  = _dernier_numero(code_str)
                poids   = _parse_montant(cell_l.value)
                cur_proj = Projet(
                    programme_id=cur_prog.id,
                    numero=numero,
                    nom=desig or f'Projet {code_str}',
                    poids=poids,
                )
                db.session.add(cur_proj)
                db.session.flush()
                cur_act = None
                nb_proj += 1
                print(f"    PROJ {code_str}: {desig[:60]}")
                continue

            # ── ACTIVITÉ ───────────────────────────────────────────────────────
            if niveau == 3:
                if cur_proj is None:
                    nb_skip += 1
                    continue
                numero    = _dernier_numero(code_str)
                poids     = _parse_montant(cell_l.value)
                imputation = str(cell_d.value).strip() if cell_d.value else None
                if imputation and imputation.lower() in ('néant', 'neant', '-', ''):
                    imputation = None
                rp = _parse_montant(cell_e.value)
                fa = _parse_montant(cell_f.value)
                fn = _parse_montant(cell_g.value)
                ap = _parse_montant(cell_h.value)
                af = _parse_montant(cell_i.value)
                debut, fin = _parse_periode(cell_k.value)
                mode = str(cell_o.value).strip() if cell_o.value else 'Direct'
                if not mode:
                    mode = 'Direct'
                svc_resp_a, dir_resp_a, unrec_m = _parse_responsable(cell_m.value)
                if svc_resp_a and not dir_resp_a:
                    dir_resp_a = svc_resp_a.direction
                svcs_ass, dirs_ass, unrec_n = _parse_structures(cell_n.value)
                tous_unrec = [t for t in unrec_m + unrec_n if t]
                acteurs_txt = ', '.join(tous_unrec) if tous_unrec else None

                cur_act = Activite(
                    projet_id=cur_proj.id,
                    numero=numero,
                    nom=desig or f'Activité {code_str}',
                    poids=poids,
                    imputation_budgetaire=imputation,
                    ressources_propres=rp,
                    fadec_affecte=fa,
                    fadec_non_affecte=fn,
                    autres_partenaires=ap,
                    autres_fonds=af,
                    periode_debut=debut,
                    periode_fin=fin,
                    mode_execution=mode,
                    direction_responsable_id=dir_resp_a.id if dir_resp_a else None,
                    acteurs_externes=acteurs_txt,
                )
                db.session.add(cur_act)
                db.session.flush()
                if svcs_ass:
                    cur_act.services_intervenants = svcs_ass
                if dirs_ass:
                    cur_act.directions_associees = dirs_ass
                nb_act += 1
                continue

            # ── TÂCHE ──────────────────────────────────────────────────────────
            if niveau == 4:
                if cur_act is None:
                    nb_skip += 1
                    continue
                numero    = _dernier_numero(code_str)
                poids     = _parse_montant(cell_l.value)
                imputation = str(cell_d.value).strip() if cell_d.value else None
                if imputation and imputation.lower() in ('néant', 'neant', '-', ''):
                    imputation = None
                rp = _parse_montant(cell_e.value)
                fa = _parse_montant(cell_f.value)
                fn = _parse_montant(cell_g.value)
                ap = _parse_montant(cell_h.value)
                af = _parse_montant(cell_i.value)
                debut, fin = _parse_periode(cell_k.value)
                mode = str(cell_o.value).strip() if cell_o.value else 'Direct'
                if not mode:
                    mode = 'Direct'
                svc_resp, dir_resp_t, unrec_m = _parse_responsable(cell_m.value)
                svcs_ass, dirs_ass, unrec_n = _parse_structures(cell_n.value)
                tous_unrec = [t for t in unrec_m + unrec_n if t]
                acteurs_txt = ', '.join(tous_unrec) if tous_unrec else None

                tache = Tache(
                    activite_id=cur_act.id,
                    numero=numero,
                    ordre=numero,
                    nom=desig or f'Tâche {code_str}',
                    poids=poids,
                    imputation_budgetaire=imputation,
                    ressources_propres=rp,
                    fadec_affecte=fa,
                    fadec_non_affecte=fn,
                    autres_partenaires=ap,
                    autres_fonds=af,
                    mode_execution=mode,
                    periode_debut=debut,
                    periode_fin=fin,
                    service_responsable_id=svc_resp.id if svc_resp else None,
                    direction_responsable_id=dir_resp_t.id if dir_resp_t else None,
                    acteurs_externes=acteurs_txt,
                )
                db.session.add(tache)
                db.session.flush()
                if svcs_ass:
                    tache.services_concernes = svcs_ass
                if dirs_ass:
                    tache.directions_associees = dirs_ass
                nb_tach += 1
                continue

            nb_skip += 1

        db.session.commit()
        print()
        print("=" * 55)
        print(f"Import terminé — Programme {prog_cible} du PTA {annee_val}")
        print(f"  Programmes : {nb_prog}")
        print(f"  Projets    : {nb_proj}")
        print(f"  Activités  : {nb_act}")
        print(f"  Tâches     : {nb_tach}")
        print(f"  Ignorées   : {nb_skip}")
        print("=" * 55)


# ─── Point d'entrée ───────────────────────────────────────────────────────────

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage : python import_programme.py <numero_programme> [chemin_excel] [annee]")
        print("Exemple : python import_programme.py 3")
        sys.exit(1)

    prog_cible = int(sys.argv[1])
    path  = sys.argv[2] if len(sys.argv) > 2 else EXCEL_PATH
    annee = int(sys.argv[3]) if len(sys.argv) > 3 else ANNEE_CIBLE

    if not os.path.exists(path):
        print(f"ERREUR : Fichier introuvable : {path}")
        sys.exit(1)

    run_import(prog_cible, path, annee)
